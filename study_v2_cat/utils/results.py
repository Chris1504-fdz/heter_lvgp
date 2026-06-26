"""
results.py — StudyResults: load the heteroscedastic LVGP BO sweep and plot it.

    from utils import StudyResults
    study = StudyResults.load("results")   # reads results/**/*.mat
    study.summary()
    study.plot_convergence()               # every plot_* returns a Figure (renders inline)

Run with the ml_gp_env Python (numpy / scipy / matplotlib).
"""
import os, glob, warnings
import numpy as np
import scipy.io
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D

from . import problem as _problem
from .problem import (LB, UB, VAR_FCTR, NOISE_MULS, CONFIG_ORDER,
                      f_true, sigma, label, acf_tag, canon_cfg, style_for,
                      ground_truth_min, true_opt_location, true_min_per_category)


def _s(x):
    a = np.ravel(x)
    return a[0] if a.size else None


# Data-source tags — stamped on every objective-value plot so it is unambiguous
# whether the y-values come from the noisy BO observations or the true function.
SRC_NOISY = "values = noisy sample-mean of f (what BO actually observes)"
SRC_TRUE  = "values = TRUE noise-free f evaluated at the design (ground truth)"
# Selection tags — for plots where is_ground_truth changes WHICH design is "best".
SEL_NOISY = "best = ranked by noisy sample-mean (what BO selects on)"
SEL_TRUE  = "best = ranked by TRUE noise-free f (ground truth)"


def _src_tag(is_ground_truth):
    return SRC_TRUE if is_ground_truth else SRC_NOISY


def _sel_tag(is_ground_truth):
    return SEL_TRUE if is_ground_truth else SEL_NOISY


class StudyResults:
    """Container for the loaded sweep runs with all study plots as methods."""

    def __init__(self, runs, results_dir):
        self.runs = runs
        self.results_dir = results_dir
        self.var_fctr = VAR_FCTR
        self._colors = None
        self._cfgs = None

    # ============================ loading ============================
    @classmethod
    def load(cls, results_dir="results"):
        """Load every run under results_dir into a list of run dicts. Reads BOTH .npz
        (this per-category-GP study) and .mat (the study_v2 LVGP study) via _load_run_file,
        so the same class plots either study and compare_studies() can overlay them."""
        runs = []
        files = (glob.glob(os.path.join(results_dir, "**", "*.npz"), recursive=True)
                 + glob.glob(os.path.join(results_dir, "**", "*.mat"), recursive=True))
        for f in sorted(files):
            r = _load_run_file(f)
            if r is not None:
                runs.append(r)
        return cls(runs, results_dir)

    # ============================ helpers ============================
    @staticmethod
    def cfg_key(r):
        p = r["param"]
        return (r["acf"], "na" if p != p else round(float(p), 6))

    def _color_map(self):
        """(colors, cfgs) for the configs present in the runs (cached). Colors come
        from the journal style (one colorblind-safe color per acquisition family)."""
        if self._colors is None:
            cfgs = sorted({self.cfg_key(r) for r in self.runs}, key=lambda c: (c[0], str(c[1])))
            self._colors = {c: self._style(c)["color"] for c in cfgs}
            self._cfgs = cfgs
        return self._colors, self._cfgs

    @staticmethod
    def _style(cfg):
        """Journal style dict {color, marker, linestyle} for a canonical cfg key."""
        param = float("nan") if cfg[1] == "na" else cfg[1]
        return style_for(cfg[0], param)

    def set_palette(self, palette):
        """Switch the per-family color palette (a name in `utils.problem.PALETTES` —
        okabe_ito, tol_bright, tol_muted, tol_vibrant, dark2, grayscale — or a list of
        hex colors). Returns self; re-run the plot cells to see the new colors."""
        _problem.set_palette(palette)
        self._colors = None; self._cfgs = None       # invalidate cached color map
        return self

    @staticmethod
    def preview_palettes():
        """Swatch chart of every named palette (rows) across the 6 families (cols)."""
        names = list(_problem.PALETTES); fams = _problem.FAMILY_ORDER
        fig, ax = plt.subplots(figsize=(1.25*len(fams)+2.5, 0.55*len(names)+1.2))
        for r, name in enumerate(names):
            y = len(names)-1-r
            for c, col in enumerate(_problem.PALETTES[name]):
                ax.add_patch(plt.Rectangle((c, y), 0.92, 0.85, color=col))
            ax.text(-0.15, y+0.42, name, ha="right", va="center", fontsize=9, family="monospace")
        for c, fam in enumerate(fams):
            ax.text(c+0.46, len(names)+0.05, fam.upper(), ha="center", va="bottom", fontsize=8)
        ax.set_xlim(-3.0, len(fams)); ax.set_ylim(0, len(names)+0.7); ax.axis("off")
        ax.set_title("Categorical palettes  (color = acquisition family; marker+linestyle stay fixed)",
                     fontsize=11)
        fig.tight_layout()
        return fig

    @staticmethod
    def _facet_legend(fig, ax, rect_right=0.85):
        """Put a single shared legend to the right of a faceted (iteration-faceted)
        figure, and pin the x-axis to start at iteration 1 across all panels."""
        for a in fig.axes:
            a.set_xlim(left=0.5)                 # iteration axis always starts at 1
        h, l = ax.get_legend_handles_labels()
        fig.legend(h, l, loc="center left", bbox_to_anchor=(rect_right + 0.01, 0.5),
                   fontsize=8, frameon=False, handlelength=2.6)
        fig.tight_layout(rect=[0, 0, rect_right, 1])

    @property
    def n_reps(self):
        return sorted({r["n_rep"] for r in self.runs})

    def filter(self, n_rep=None, acf=None):
        """Subset of runs matching n_rep and/or acf (returns a list of run dicts)."""
        out = self.runs
        if n_rep is not None:
            out = [r for r in out if r["n_rep"] == n_rep]
        if acf is not None:
            out = [r for r in out if r["acf"] == acf]
        return out

    def ground_truth_min(self):
        return ground_truth_min()

    def best_obj_per_category(self, run):
        """Lowest NOISY sample-mean objective found in each category level (NaN if none)."""
        X, Y = run["X_sampled"], run["Y_sampled"]
        lv = np.round(X[:, 1]).astype(int)
        out = np.full(len(VAR_FCTR), np.nan)
        for L in range(1, len(VAR_FCTR)+1):
            m = lv == L
            if m.any():
                out[L-1] = Y[m].min()
        return out

    @staticmethod
    def _true_obj_of_samples(run):
        """True noise-free f at every sampled design (x1, level) of the run."""
        X = run["X_sampled"]; lv = np.round(X[:, 1]).astype(int)
        out = np.full(len(X), np.nan)
        for i in range(len(X)):
            if 1 <= lv[i] <= len(VAR_FCTR):
                out[i] = float(f_true(np.array([X[i, 0]]), VAR_FCTR[lv[i]-1])[0])
        return out

    @staticmethod
    def _true_var_of_samples(run):
        """Analytic TRUE noise variance σ²(x1, level) at every sampled design."""
        X = run["X_sampled"]; lv = np.round(X[:, 1]).astype(int)
        out = np.full(len(X), np.nan)
        for i in range(len(X)):
            if 1 <= lv[i] <= len(VAR_FCTR):
                out[i] = float(sigma(X[i, 0], lv[i])**2)
        return out

    def _best_sample_idx(self, run, is_ground_truth):
        """Index of the 'best' sampled design — ranked by the TRUE f
        (is_ground_truth=True) or by the NOISY sample-mean Y_sampled (False)."""
        key = self._true_obj_of_samples(run) if is_ground_truth else run["Y_sampled"]
        return int(np.nanargmin(key))

    def true_obj_per_category(self, run):
        """Best TRUE noise-free objective among designs sampled in each category
        (selected AND evaluated by f — the ground-truth best per level)."""
        lv = np.round(run["X_sampled"][:, 1]).astype(int)
        ft = self._true_obj_of_samples(run)
        out = np.full(len(VAR_FCTR), np.nan)
        for L in range(1, len(VAR_FCTR)+1):
            m = lv == L
            if m.any():
                out[L-1] = np.nanmin(ft[m])
        return out

    def _final_obj(self, run, is_ground_truth=False):
        """Objective of the 'best' sampled design. is_ground_truth=True: TRUE f at the design
        with the LOWEST TRUE f among all samples (best-by-ground-truth -- matches the
        true_best_sampled convergence metric; never uses the noisy mean). False: the best
        OBSERVED noisy sample-mean (the design BO would recommend)."""
        idx = self._best_sample_idx(run, is_ground_truth)
        return float(self._true_obj_of_samples(run)[idx] if is_ground_truth else run["Y_sampled"][idx])

    def _final_var(self, run, is_ground_truth=False):
        """Aleatoric variance of the 'best' sampled design (selected exactly as in _final_obj):
        is_ground_truth=True -> analytic TRUE σ²(x1,level) at the best-by-true design; False ->
        the n_rep sample-variance estimate at the best-by-noisy design."""
        idx = self._best_sample_idx(run, is_ground_truth)
        return float(self._true_var_of_samples(run)[idx] if is_ground_truth else run["Y_var_sampled"][idx])

    def summary(self):
        """Quick text summary of what's loaded."""
        cfgs = sorted({self.cfg_key(r) for r in self.runs})
        print(f"{len(self.runs)} runs | {len(cfgs)} configs | "
              f"n_rep={self.n_reps} | seeds={len({r['seed'] for r in self.runs})} | "
              f"ground-truth min = {ground_truth_min():.3f}")

    # ============================ overview ============================
    def plot_progress(self, target_per_acq=90):
        """One bar per acquisition (count of completed .mat) with a target line.
        Reads the results directory directly, so it also works mid-sweep."""
        labels, counts = [], []
        for acf, p in CONFIG_ORDER:
            d = os.path.join(self.results_dir, acf_tag(acf, p))
            counts.append(sum(len(glob.glob(os.path.join(d, "**", e), recursive=True))
                              for e in ("*.npz", "*.mat")))   # this study saves .npz
            labels.append(label(acf, p))
        counts = np.array(counts)
        done, total = int(counts.sum()), target_per_acq * len(CONFIG_ORDER)
        fig, ax = plt.subplots(figsize=(11, 5))
        colors = ["#2e7d32" if c >= target_per_acq else "#4a90d9" for c in counts]
        bars = ax.bar(range(len(counts)), counts, color=colors, edgecolor="k", lw=0.4)
        ax.axhline(target_per_acq, color="crimson", ls="--", lw=1.5, label=f"target = {target_per_acq}")
        for b, c in zip(bars, counts):
            ax.text(b.get_x()+b.get_width()/2, c+target_per_acq*0.015, str(int(c)),
                    ha="center", va="bottom", fontsize=8)
        ax.set_xticks(range(len(labels))); ax.set_xticklabels(labels, rotation=40, ha="right", fontsize=8)
        ax.set_ylim(0, target_per_acq*1.12); ax.set_ylabel("completed runs")
        pct = 100*done/total if total else 0
        ax.set_title(f"Sweep progress — {done}/{total} runs ({pct:.0f}%)  "
                     f"·  {int((counts >= target_per_acq).sum())}/{len(counts)} configs complete")
        ax.legend(loc="lower right"); ax.grid(axis="y", alpha=0.3); fig.tight_layout()
        return fig

    def plot_initial_doe(self, n_rep=None):
        """Initial DOE (LHS starting points, before BO) overlaid across seeds,
        x1 vs category. DOE is identical across acquisitions for a (seed, n_rep),
        so it is deduplicated. n_rep=None pools all settings."""
        gt_lv, gt_x1 = true_opt_location()
        seen = {}
        for r in self.runs:
            if n_rep is not None and r["n_rep"] != n_rep:
                continue
            seen.setdefault((r["seed"], r["n_rep"]), r["X_sampled"][:r["n_initial"]])
        if not seen:
            raise ValueError("no runs match")
        pts = np.vstack(list(seen.values()))
        lv = np.round(pts[:, 1]).astype(int); nlev = len(VAR_FCTR)
        fig, ax = plt.subplots(figsize=(8, 5)); rng = np.random.default_rng(0)
        ax.scatter(lv + rng.uniform(-0.18, 0.18, len(lv)), pts[:, 0], s=16, c="steelblue",
                   alpha=0.35, edgecolor="k", linewidth=0.2, label="initial DOE points")
        ax.plot(gt_lv, gt_x1, marker="*", ms=16, c="crimson", zorder=5, label="true optimum")
        ax.set_xticks(range(1, nlev+1))
        ax.set_xticklabels([f"{i+1}\n({v:g})" for i, v in enumerate(VAR_FCTR)])
        ax.set_xlabel("categorical level (value)"); ax.set_ylabel("$x_1$")
        ax.set_ylim(LB-0.5, UB+0.5)
        per_cat = len(pts) // (len(seen) * nlev)
        nrtxt = "all n_rep" if n_rep is None else f"n_rep={n_rep}"
        ax.set_title(f"Initial DOE across {len(seen)} designs ({nrtxt}) — "
                     f"{len(pts)} points ({per_cat} LHS per category)")
        ax.grid(alpha=0.3); ax.legend(loc="upper right"); fig.tight_layout()
        return fig

    # ============================ convergence ============================
    def plot_convergence(self, gt=None):
        """Best sample-mean objective vs iteration, mean ± 95% CI, faceted by n_rep."""
        if gt is None: gt = ground_truth_min()
        colors, cfgs = self._color_map()
        n_reps = self.n_reps
        fig, axes = plt.subplots(1, len(n_reps), figsize=(6*len(n_reps), 4.5),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            for i, cfg in enumerate(cfgs):
                hist = [r["Y_min_history"] for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if not hist: continue
                L = min(len(h) for h in hist); H = np.array([h[:L] for h in hist])
                mean = H.mean(0)
                sem = H.std(0, ddof=1)/np.sqrt(H.shape[0]) if H.shape[0] > 1 else np.zeros(L)
                it = np.arange(1, L+1); st = self._style(cfg); me = max(1, L//6)
                ax.plot(it, mean, color=st["color"], ls=st["linestyle"], marker=st["marker"],
                        markevery=me, ms=4.5, lw=1.6, label=label(*cfg))
                ax.fill_between(it, mean-1.96*sem, mean+1.96*sem, color=st["color"], alpha=0.12)
            ax.axhline(gt, ls=(0, (6, 4)), color="0.3", lw=1.2, label="ground truth")
            ax.set_title(f"n_rep = {nr}"); ax.set_xlabel("Iteration"); ax.grid(alpha=0.3)
        axes[0][0].set_ylabel("Best sample-mean objective $y^*$")
        fig.suptitle(f"Best objective vs iteration (mean ± 95% CI)\n{SRC_NOISY}")
        self._facet_legend(fig, axes[0][0])
        return fig

    @staticmethod
    def _true_obj_at_recommended(run):
        out = []
        for xr in run["X_min_est"]:
            lv = int(round(xr[1]))
            out.append(float(f_true(np.array([xr[0]]), VAR_FCTR[lv-1])[0])
                       if 1 <= lv <= len(VAR_FCTR) else np.nan)
        return np.array(out)

    def plot_convergence_true(self, gt=None, log=False, ymax=None):
        """Convergence using the TRUE noise-free objective at the recommended
        optimum (X_min_est) each iteration — sits ABOVE the true global min.

        The early iterations spike (~30-40) and squash the convergence band on a
        linear axis. Use log=True (log y-axis, recommended — values are always
        >= the true min > 0) and/or ymax=<v> to zoom into the convergence region."""
        if gt is None: gt = ground_truth_min()
        colors, cfgs = self._color_map()
        n_reps = self.n_reps
        fig, axes = plt.subplots(1, len(n_reps), figsize=(6*len(n_reps), 4.5),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            for i, cfg in enumerate(cfgs):
                curves = [self._true_obj_at_recommended(r) for r in self.runs
                          if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if not curves: continue
                L = min(len(c) for c in curves); C = np.array([c[:L] for c in curves])
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    mean = np.nanmean(C, 0)
                    sem = (np.nanstd(C, 0, ddof=1)/np.sqrt(C.shape[0])) if C.shape[0] > 1 else np.zeros(L)
                sem = np.nan_to_num(sem); it = np.arange(1, L+1); st = self._style(cfg); me = max(1, L//6)
                ax.plot(it, mean, color=st["color"], ls=st["linestyle"], marker=st["marker"],
                        markevery=me, ms=4.5, lw=1.6, label=label(*cfg))
                lo = mean - 1.96*sem
                if log: lo = np.maximum(lo, gt*0.5)          # keep CI band positive for log
                ax.fill_between(it, lo, mean+1.96*sem, color=st["color"], alpha=0.12)
            ax.axhline(gt, ls=(0, (6, 4)), color="0.3", lw=1.2, label="true global min")
            ax.set_title(f"n_rep = {nr}"); ax.set_xlabel("Iteration"); ax.grid(alpha=0.3, which="both")
            if log:
                ax.set_yscale("log"); ax.set_ylim(bottom=max(gt*0.7, 0.3))
            if ymax is not None:
                ax.set_ylim(top=ymax)
        scale = " — log scale" if log else (f" — zoom (y≤{ymax:g})" if ymax is not None else "")
        axes[0][0].set_ylabel("True objective at recommended optimum")
        fig.suptitle(f"Convergence vs ground truth — true obj. at recommended optimum (mean ± 95% CI){scale}\n"
                     f"{SRC_TRUE} (at X_min_est)")
        self._facet_legend(fig, axes[0][0])
        return fig

    def plot_simple_regret(self, gt=None, is_ground_truth=False, eps=None):
        """Simple regret |best − truth| vs iteration (log scale) — convergence rate.
        is_ground_truth=False: regret of the NOISY best sample-mean (Y_min_history).
        is_ground_truth=True:  regret of the TRUE f at the recommended optimum (X_min_est).
        On a log axis regret -> 0 is the (unreachable) bottom, so instead of a 0-line
        pass eps=<tol> (e.g. 0.1) to draw an ε-accuracy line — read off how many
        iterations each method needs to get within ε of the truth (cf. RAHBO Cor. 1.1)."""
        if gt is None: gt = ground_truth_min()
        colors, cfgs = self._color_map()
        n_reps = self.n_reps
        fig, axes = plt.subplots(1, len(n_reps), figsize=(6*len(n_reps), 4.5),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            for i, cfg in enumerate(cfgs):
                if is_ground_truth:
                    series = [self._true_obj_at_recommended(r) for r in self.runs
                              if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                else:
                    series = [r["Y_min_history"] for r in self.runs
                              if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if not series: continue
                L = min(len(h) for h in series); H = np.array([h[:L] for h in series])
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    regret = np.nanmean(np.maximum(np.abs(H - gt), 1e-6), 0)
                st = self._style(cfg); me = max(1, L//6)
                ax.semilogy(np.arange(1, L+1), regret, color=st["color"], ls=st["linestyle"],
                            marker=st["marker"], markevery=me, ms=4.5, lw=1.6, label=label(*cfg))
            if eps is not None:
                ax.axhline(eps, ls=(0, (6, 4)), color="0.3", lw=1.2, label=f"ε = {eps:g}")
            ax.set_title(f"n_rep = {nr}"); ax.set_xlabel("Iteration"); ax.grid(alpha=0.3, which="both")
        axes[0][0].set_ylabel("mean |best − truth|  (log)")
        fig.suptitle(f"Simple regret vs iteration (convergence rate)\n{_src_tag(is_ground_truth)}")
        self._facet_legend(fig, axes[0][0])
        return fig

    def regret_table(self, gt=None, is_ground_truth=True, eps=0.1, excel_path=None):
        """Per-config × n_rep summary of the simple-regret plot, as a pandas table.

        For each acquisition config and n_rep it reports the FINAL simple regret
        |f − f*| of the recommended optimum, summarised over the 30 seeds:
          mean   — the curve the plot draws (tail/outlier-sensitive on a log axis),
          median — robust centre (shows when a high mean is just a few bad seeds),
          succ%  — fraction of seeds whose final recommendation is within eps of f*,
          it<=eps— first iteration at which the MEAN regret drops to <= eps ('—' if never).
        Columns are a (n_rep, metric) MultiIndex; rows are the 12 configs grouped by
        family. is_ground_truth=True (default) matches plot_simple_regret's TRUE curve.
        Pass excel_path to also write a styled .xlsx. Returns the DataFrame."""
        import pandas as pd
        if gt is None: gt = ground_truth_min()
        n_reps = self.n_reps
        cfgs = [canon_cfg(a, p) for a, p in CONFIG_ORDER]
        fams = {canon_cfg(a, p): ("risk-neutral" if a in ("lcb", "pi", "ei")
                                  else "risk-aware") for a, p in CONFIG_ORDER}
        rows = {}
        for a, p in CONFIG_ORDER:
            cfg = canon_cfg(a, p)
            rec = {("", "family"): fams[cfg]}
            for nr in n_reps:
                series = [self._true_obj_at_recommended(r) if is_ground_truth
                          else r["Y_min_history"] for r in self.runs
                          if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if not series:
                    for m in ("mean", "median", "succ%", "it<=eps"):
                        rec[(f"n_rep={nr}", m)] = np.nan
                    continue
                L = min(len(h) for h in series); H = np.array([h[:L] for h in series])
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    reg = np.nanmean(np.maximum(np.abs(H - gt), 1e-6), 0)
                finals = np.abs(np.array([h[-1] for h in series]) - gt)
                below = np.where(reg <= eps)[0]
                rec[(f"n_rep={nr}", "mean")]    = float(np.mean(finals))
                rec[(f"n_rep={nr}", "median")]  = float(np.median(finals))
                rec[(f"n_rep={nr}", "succ%")]   = round(100 * np.mean(finals <= eps))
                rec[(f"n_rep={nr}", "it<=eps")] = int(below[0] + 1) if len(below) else np.nan
            rows[label(a, p)] = rec
        df = pd.DataFrame(rows).T
        df.columns = pd.MultiIndex.from_tuples(df.columns)
        df.index.name = f"config (eps={eps:g}, f*={gt:.3f})"
        if excel_path:
            src = "TRUE f at recommended optimum" if is_ground_truth else "noisy sample-mean"
            note = (f"final simple regret |f-f*|, src={src}, eps={eps:g}, f*={gt:.4f}; "
                    f"mean/median/succ% over 30 seeds; it<=eps = first iter mean regret <= eps")
            self._write_table_excel(df, excel_path, "simple_regret",
                                    {"mean", "median"}, {"succ%"}, note)
            print("wrote", excel_path)
        return df

    @staticmethod
    def _write_table_excel(df, path, sheet, lowbetter, pctcols, note):
        """Write a (config × n_rep) summary DataFrame to a styled .xlsx.
        lowbetter = set of sub-column names to format 0.0000 with a green(low)->
        red(high) gradient; pctcols = sub-columns with a red(0)->green(100) gradient.
        note = an italic footnote describing the metrics."""
        import pandas as pd
        from openpyxl.styles import Font, Alignment
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.utils import get_column_letter
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            df.to_excel(xw, sheet_name=sheet)
            ws = xw.sheets[sheet]
            for cell in ws[1] + ws[2]:
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "B3"
            for j, (top, sub) in enumerate(df.columns, start=2):
                col = ws.cell(row=2, column=j).column_letter
                if sub in lowbetter:
                    for r in range(3, ws.max_row + 1):
                        ws.cell(row=r, column=j).number_format = "0.0000"
                    ws.conditional_formatting.add(
                        f"{col}3:{col}{ws.max_row}",
                        ColorScaleRule(start_type="min", start_color="63BE7B",
                                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                                       end_type="max", end_color="F8696B"))
                elif sub in pctcols:
                    ws.conditional_formatting.add(
                        f"{col}3:{col}{ws.max_row}",
                        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                                       end_type="num", end_value=100, end_color="63BE7B"))
            for ci in range(1, ws.max_column + 1):
                letter = get_column_letter(ci)
                w = max((len(str(ws.cell(row=r, column=ci).value))
                         for r in range(1, ws.max_row + 1)
                         if ws.cell(row=r, column=ci).value is not None), default=0)
                ws.column_dimensions[letter].width = max(9, w + 2)
            cell = ws.cell(row=ws.max_row + 2, column=1, value=note)
            cell.font = Font(italic=True, size=9)

    # ============================ x1 location ============================
    def _x1_curve(self, run, source, is_ground_truth):
        """(x1, level) of the best-so-far design per iteration. is_ground_truth ranks
        the cumulative best by TRUE f; else by the model recommendation (source=
        'recommended', X_min_est) or the noisy best sample (source='best')."""
        if is_ground_truth:
            key = self._true_obj_of_samples(run); X = run["X_sampled"]; n0 = run["n_initial"]
            idx = [int(np.nanargmin(key[:n0+i])) for i in range(1, len(key)-n0+1)]
            return X[idx, 0], np.round(X[idx, 1]).astype(int)
        if source == "best":
            X, Y, n0 = run["X_sampled"], run["Y_sampled"], run["n_initial"]
            idx = [int(np.argmin(Y[:n0+i])) for i in range(1, len(Y)-n0+1)]
            return X[idx, 0], np.round(X[idx, 1]).astype(int)
        Xme = run["X_min_est"]
        return Xme[:, 0], np.round(Xme[:, 1]).astype(int)

    def _x1_final(self, run, source, is_ground_truth):
        if is_ground_truth:
            return float(run["X_sampled"][self._best_sample_idx(run, True), 0])
        return float(run["X_best_final"][0]) if source == "best" else float(run["X_min_est"][-1, 0])

    @staticmethod
    def _x1_src_label(source, is_ground_truth):
        if is_ground_truth: return "ground-truth best sample"
        return "noisy best sample" if source == "best" else "model recommendation"

    def plot_x1_convergence(self, source="recommended", is_ground_truth=False,
                            show_off_category=False):
        """x1 of the best location vs iteration (mean ± 95% CI) vs the true optimal x1.
        is_ground_truth=True ranks the running best by the TRUE f; otherwise source=
        'recommended' (X_min_est) or 'best' (noisy best-by-sample-mean).
        show_off_category=False (default) masks iterations whose best design is NOT in
        category 2 (so the curve only averages comparable category-2 x1); set True to
        include every iteration's x1 regardless of category (continuous from iter 1, but
        mixes x1 across categories)."""
        gt_lv, gt_x1 = true_opt_location()
        colors, cfgs = self._color_map()
        n_reps = self.n_reps
        fig, axes = plt.subplots(1, len(n_reps), figsize=(6*len(n_reps), 4.5),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            for i, cfg in enumerate(cfgs):
                curves = []
                for r in self.runs:
                    if self.cfg_key(r) != cfg or r["n_rep"] != nr: continue
                    x1c, lvc = self._x1_curve(r, source, is_ground_truth)
                    curves.append(x1c if show_off_category
                                  else np.where(lvc == gt_lv, x1c, np.nan))   # mask off-category
                if not curves: continue
                L = min(len(c) for c in curves); C = np.array([c[:L] for c in curves])
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    mean = np.nanmean(C, 0)
                    sem = (np.nanstd(C, 0, ddof=1)/np.sqrt(np.maximum(np.sum(~np.isnan(C), 0), 1)))
                it = np.arange(1, L+1); st = self._style(cfg); me = max(1, L//6)
                ax.plot(it, mean, color=st["color"], ls=st["linestyle"], marker=st["marker"],
                        markevery=me, ms=4.5, lw=1.6, label=label(*cfg))
                ax.fill_between(it, mean-1.96*np.nan_to_num(sem), mean+1.96*np.nan_to_num(sem),
                                color=st["color"], alpha=0.12)
            ax.axhline(gt_x1, ls=(0, (6, 4)), color="0.3", lw=1.4, label=f"true $x_1$ = {gt_x1:.2f}")
            ax.set_title(f"n_rep = {nr}"); ax.set_xlabel("Iteration"); ax.grid(alpha=0.3)
        src = self._x1_src_label(source, is_ground_truth)
        catnote = "any category" if show_off_category else "category 2 only"
        axes[0][0].set_ylabel(f"$x_1$ of {src} ({catnote})")
        masknote = ("off-category iterations INCLUDED" if show_off_category
                    else "off-category iterations masked")
        fig.suptitle(f"Best $x_1$ vs iteration — convergence to true optimum $x_1$ ({masknote})\n"
                     + _sel_tag(is_ground_truth))
        self._facet_legend(fig, axes[0][0])
        return fig

    def plot_x1_distribution(self, n_rep=10, source="recommended", show_violin=True,
                             show_mean=True, is_ground_truth=False, show_off_category=False):
        """Distribution of the FINAL best x1 across seeds, per acquisition, vs the true
        optimal x1 (category 2). Reads final_design_table, so ONLY designs whose best is
        actually in category 2 enter the distribution; the count of seeds whose best is
        in another category is shown under each acquisition label. show_off_category=True
        overlays those other-category x1 as hollow diamonds."""
        gt_lv, gt_x1 = true_opt_location()
        colors, cfgs = self._color_map()
        table = self.final_design_table(source, is_ground_truth)
        labs, cols, in2, offx = [], [], [], []
        for cfg in cfgs:
            recs = [t for t in table if t["cfg"] == cfg and t["n_rep"] == n_rep]
            if not recs:
                continue
            x2 = [t["x1"] for t in recs if t["in_cat2"]]
            xo = [t["x1"] for t in recs if not t["in_cat2"]]
            in2.append(x2); offx.append(xo); cols.append(colors[cfg])
            labs.append(label(*cfg) + (f"\n({len(xo)} off-cat)" if xo else ""))
        fig, ax = plt.subplots(figsize=(1.1*len(labs)+3, 5)); pos = np.arange(1, len(labs)+1)
        rng = np.random.default_rng(0)
        for i, d, c in zip(pos, in2, cols):
            if show_violin and len(d) >= 2:
                parts = ax.violinplot([d], positions=[i], showmeans=False, showextrema=False, widths=0.8)
                for pc in parts["bodies"]:
                    pc.set_facecolor(c); pc.set_alpha(0.35)
            if d:
                ax.scatter(i + rng.uniform(-0.1, 0.1, len(d)), d, s=16, color=c,
                           alpha=0.7, edgecolor="k", linewidth=0.25, zorder=3)
                if show_mean:
                    ax.plot([i-0.28, i+0.28], [np.mean(d)]*2, color="k", lw=2, zorder=4)
        if show_off_category:
            for i, xo in zip(pos, offx):
                if xo:
                    ax.scatter(i + rng.uniform(-0.1, 0.1, len(xo)), xo, s=24, facecolors="none",
                               edgecolors="0.25", marker="D", linewidth=1.0, zorder=4)
        ax.axhline(gt_x1, ls="--", color="k", lw=1.6, label=f"true $x_1$ = {gt_x1:.2f}")
        ax.set_xticks(pos); ax.set_xticklabels(labs, rotation=40, ha="right", fontsize=8)
        src = self._x1_src_label(source, is_ground_truth)
        ax.set_ylabel(f"final $x_1$ of {src} (category 2 only)")
        ax.set_title(f"Distribution of final best $x_1$ vs true optimum (n_rep={n_rep}, category 2)\n"
                     f"{_sel_tag(is_ground_truth)}")
        ax.grid(alpha=0.3, axis="y"); ax.legend(loc="upper right", fontsize=9); fig.tight_layout()
        return fig

    def _final_design(self, run, source, is_ground_truth):
        """(x1, level) of the reported/best final design used by the x1 plots."""
        if is_ground_truth:
            xr = run["X_sampled"][self._best_sample_idx(run, True)]
        elif source == "best":
            xr = run["X_best_final"]
        else:
            xr = run["X_min_est"][-1]
        return float(xr[0]), int(round(xr[1]))

    def final_design_table(self, source="recommended", is_ground_truth=False):
        """Single source of truth for the x1-location plots: a per-run record of each
        run's FINAL best design under the given selection, computed once and cached
        (also stored on self.final_designs for inspection). Every record carries the
        design's LEVEL, so downstream plots never blindly assume category 2.

        record = dict(cfg, acf, param, n_rep, seed, x1, level, in_cat2,
                      obj_true  = true noise-free f in the design's OWN category,
                      obj_noisy = Y_best_final)."""
        key = (source, bool(is_ground_truth))
        cache = getattr(self, "_final_tables", None)
        if cache is None:
            cache = self._final_tables = {}
        if key not in cache:
            gt_lv = true_opt_location()[0]
            recs = []
            for r in self.runs:
                x1, lvl = self._final_design(r, source, is_ground_truth)
                recs.append(dict(
                    cfg=self.cfg_key(r), acf=r["acf"], param=r["param"],
                    n_rep=r["n_rep"], seed=r["seed"], x1=x1, level=lvl,
                    in_cat2=(lvl == gt_lv),
                    obj_true=float(f_true(np.array([x1]), VAR_FCTR[lvl-1])[0]),
                    obj_noisy=float(r["Y_best_final"])))
            cache[key] = recs
        self.final_designs = cache[key]      # expose the most-recently-built table
        return cache[key]

    def plot_x1_landing(self, n_rep=10, source="recommended", is_ground_truth=False,
                        basin_split=6.0, ymax=26, show_off_category=True):
        """Where each seed's final best design lands relative to the category-2 true
        function. One panel per acquisition: the noise-free f for the optimal category
        (level 2) + its ±1σ band; each seed's best design is plotted at its TRUE
        objective IN ITS OWN category. Designs that are actually in category 2 sit ON
        the curve (filled dots); designs whose best is in ANOTHER category are flagged
        (open diamonds) at their real (own-category) objective — so an off-category
        best no longer gets lifted onto the cat-2 wall.
        show_off_category=False hides those other-category bests (the count is still
        reported in the panel title). basin_split = x1 boundary for the 'opt basin' %;
        ymax zooms the objective axis."""
        gt_lv, gt_x1 = true_opt_location()          # (2, 3.18)
        val = VAR_FCTR[gt_lv-1]
        x1g = np.linspace(LB, UB, 400)
        fg = f_true(x1g, val); sg = sigma(x1g, gt_lv)
        fopt = float(f_true(np.array([gt_x1]), val)[0])
        colors, _ = self._color_map()
        cfgs = [canon_cfg(a, p) for (a, p) in CONFIG_ORDER]
        cols = 4; rows = int(np.ceil(len(cfgs)/cols)); ylo = -3
        fig, axes = plt.subplots(rows, cols, figsize=(3.7*cols, 2.9*rows),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, cfg in zip(axes.ravel(), cfgs):
            ax.fill_between(x1g, fg-sg, fg+sg, color="0.85", zorder=0)
            ax.plot(x1g, fg, "k-", lw=1.2, zorder=1)
            ax.axvline(basin_split, color="0.6", ls=":", lw=1, zorder=1)
            recs = [t for t in self.final_design_table(source, is_ground_truth)
                    if t["cfg"] == cfg and t["n_rep"] == n_rep]
            if recs:
                x1s = np.array([t["x1"] for t in recs]); yown = np.array([t["obj_true"] for t in recs])
                in2 = np.array([t["in_cat2"] for t in recs])
                ax.scatter(x1s[in2], yown[in2], color=colors.get(cfg, "gray"), s=30, alpha=0.6,
                           edgecolor="k", linewidth=0.3, zorder=3, label="best in cat 2")
                if show_off_category and (~in2).any():
                    ax.scatter(x1s[~in2], yown[~in2], facecolors="none", edgecolors="0.25",
                               marker="D", s=34, linewidth=1.1, zorder=4, label="best in OTHER cat")
                rug = x1s if show_off_category else x1s[in2]
                ax.plot(rug, np.full_like(rug, ylo+0.8), "|", color=colors.get(cfg, "gray"),
                        ms=7, alpha=0.5, zorder=2)            # rug for x1 density
                pct = 100*np.mean(in2 & (x1s < basin_split))
                noff = int((~in2).sum())
                extra = f", {noff} off-cat" if noff else ""
                ax.set_title(f"{label(*cfg)} — {pct:.0f}% opt basin{extra}", fontsize=9)
            else:
                ax.set_title(f"{label(*cfg)}", fontsize=9, color="gray")
            ax.axhline(fopt, color="crimson", ls="--", lw=1.2, zorder=5)   # true global min level
            ax.set_ylim(ylo, ymax); ax.grid(alpha=0.3)
            # ax.set_ylim(0, 1); ax.grid(alpha=0.3)
        for ax in axes.ravel()[len(cfgs):]: ax.set_visible(False)
        for ax in axes[-1]: ax.set_xlabel("$x_1$")
        for ax in axes[:, 0]: ax.set_ylabel("true objective (own cat)")
        handles = [Line2D([], [], marker="o", ls="", color="0.4", label="best in cat 2")]
        if show_off_category:
            handles.append(Line2D([], [], marker="D", ls="", mfc="none", mec="0.25",
                                  label="best in OTHER cat"))
        handles.append(Line2D([], [], color="crimson", ls="--", label=f"true global min (f={fopt:.2f})"))
        fig.legend(handles=handles, loc="lower center", ncol=len(handles), frameon=False,
                   fontsize=9, bbox_to_anchor=(0.5, -0.02))
        fig.suptitle(f"Where the final design lands vs the category-2 true function "
                     f"(n_rep={n_rep})\n{_sel_tag(is_ground_truth)}", y=1.0)
        fig.tight_layout(rect=[0, 0.03, 1, 1])
        return fig

    # ============================ input / categorical ============================
    def plot_input_space(self, n_rep=None, cmap="viridis"):
        """x1 vs categorical level sampling, colored by BO iteration, per acquisition.
        cmap = the sequential colormap used for the BO-iteration colouring."""
        colors, cfgs = self._color_map()
        if n_rep is None: n_rep = max({r["n_rep"] for r in self.runs})
        cfgs = [c for c in cfgs if any(self.cfg_key(r) == c and r["n_rep"] == n_rep for r in self.runs)]
        nlev = len(VAR_FCTR); cols = min(4, len(cfgs)); rows = int(np.ceil(len(cfgs)/cols))
        fig, axes = plt.subplots(rows, cols, figsize=(4*cols, 3.2*rows),
                                 squeeze=False, sharex=True, sharey=True)
        sc = None
        for ax, cfg in zip(axes.ravel(), cfgs):
            init_x, bo_x, bo_it = [], [], []
            for r in self.runs:
                if self.cfg_key(r) != cfg or r["n_rep"] != n_rep: continue
                X, n0 = r["X_sampled"], r["n_initial"]
                init_x.append(X[:n0]); bo_x.append(X[n0:]); bo_it.append(np.arange(1, len(X)-n0+1))
            if init_x:
                I = np.vstack(init_x); B = np.vstack(bo_x); T = np.concatenate(bo_it)
                ax.scatter(I[:, 1], I[:, 0], c="0.6", marker="x", s=25, label="initial DOE")
                sc = ax.scatter(B[:, 1], B[:, 0], c=T, cmap=cmap, s=22,
                                edgecolor="k", linewidth=0.3, label="BO samples")
            ax.set_title(label(*cfg), fontsize=9); ax.set_xticks(range(1, nlev+1))
            ax.set_xticklabels([f"{i+1}\n({v:g})" for i, v in enumerate(VAR_FCTR)], fontsize=7)
            ax.grid(alpha=0.3)
        for ax in axes[-1]: ax.set_xlabel("level idx (value)")
        for ax in axes[:, 0]: ax.set_ylabel("$x_1$")
        if sc is not None:
            fig.colorbar(sc, ax=axes.ravel().tolist(), label="BO iteration", shrink=0.6)
        fig.suptitle(f"Input-space sampling (n_rep={n_rep}, pooled over seeds)")
        return fig

    def plot_level_histogram(self):
        """Categorical-level selection frequency (BO samples), one panel per n_rep."""
        n_reps = self.n_reps; nlev = len(VAR_FCTR)
        fig, axes = plt.subplots(1, len(n_reps), figsize=(5*len(n_reps), 4),
                                 squeeze=False, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            counts = np.zeros(nlev)
            for r in self.runs:
                if r["n_rep"] != nr: continue
                for lv in r["X_sampled"][r["n_initial"]:, 1].astype(int):
                    if 1 <= lv <= nlev: counts[lv-1] += 1
            ax.bar(range(1, nlev+1), counts, color="steelblue", edgecolor="k")
            ax.set_title(f"n_rep = {nr}"); ax.set_xlabel("categorical level (value)")
            ax.set_xticks(range(1, nlev+1))
            ax.set_xticklabels([f"{i+1}\n({v:g})" for i, v in enumerate(VAR_FCTR)], fontsize=8)
            ax.grid(alpha=0.3, axis="y")
        axes[0][0].set_ylabel("# BO samples")
        fig.suptitle("Categorical level selection frequency"); fig.tight_layout()
        return fig

    # ============================ objective vs variance ============================
    def plot_objective_variance(self, is_ground_truth=False, aggregate=True, errorbars=False):
        """Final solutions in objective vs aleatoric-variance space (per acq + n_rep).
        is_ground_truth=True plots BOTH axes from the ground truth: the noise-free f and
        the analytic noise variance σ²(x1,level) at the final design; False plots what BO
        observed: Y_best_final and the n_rep sample variance.
        aggregate=True (default) plots the MEAN (objective, variance) per (acquisition,
        n_rep) -- one point each, IDENTICAL to the values in plot_summary_heatmaps, so the
        scatter and the heatmap are directly consistent (same range). aggregate=False plots
        EVERY final solution (one point per seed), where a few n_rep=3 outliers reach much
        higher σ² (the optimistic-bias effect) -- a wider spread than the mean heatmap.
        errorbars (aggregate mode only): False = none; True/'sem' = ± standard error of the
        mean over seeds; 'std' = ± standard deviation (the spread). Drawn on both axes."""
        colors, cfgs = self._color_map()
        fig, ax = plt.subplots(figsize=(7, 5.5)); markers = {3: "o", 5: "s", 10: "^"}
        for cfg in cfgs:
            for nr in self.n_reps:
                pts = [(self._final_obj(r, is_ground_truth), self._final_var(r, is_ground_truth))
                       for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if not pts: continue
                pts = np.array(pts)
                if aggregate and errorbars and len(pts) > 1:
                    mean = pts.mean(axis=0); sd = pts.std(axis=0, ddof=1)
                    err = sd / np.sqrt(len(pts)) if errorbars in (True, "sem") else sd
                    ax.errorbar(mean[0], mean[1], xerr=err[0], yerr=err[1], fmt=markers.get(nr, "o"),
                                ms=8, color=colors[cfg], ecolor=colors[cfg], elinewidth=1, capsize=3,
                                alpha=0.85, markeredgecolor="k", markeredgewidth=0.3)
                    continue
                if aggregate:
                    pts = pts.mean(axis=0, keepdims=True)     # one MEAN point per (acq, n_rep)
                ax.scatter(pts[:, 0], pts[:, 1], color=colors[cfg], marker=markers.get(nr, "o"),
                           s=(75 if aggregate else 40), alpha=0.85, edgecolor="k", linewidth=0.3)
        cfg_h = [Line2D([], [], color=colors[c], marker="o", ls="", label=label(*c)) for c in cfgs]
        nr_h = [Line2D([], [], color="0.4", marker=markers[k], ls="", label=f"n_rep={k}")
                for k in markers if any(r["n_rep"] == k for r in self.runs)]
        ax.add_artist(ax.legend(handles=cfg_h, title="acquisition", fontsize=8, loc="upper right"))
        ax.legend(handles=nr_h, title="replicates", fontsize=8, loc="lower right")
        ax.set_xlabel("TRUE objective $f(x^*)$" if is_ground_truth else "Objective of final solution  $y$ (noisy)")
        ax.set_ylabel("TRUE noise variance σ²(x*)" if is_ground_truth
                      else "Aleatoric variance (n_rep sample estimate)")
        ttl = ("Mean final solution per (acq, n_rep)" if aggregate
               else "Final solutions in objective–variance space (per seed)")
        ax.set_title(f"{ttl}\n{_src_tag(is_ground_truth)}")
        ax.grid(alpha=0.3); fig.tight_layout()
        return fig

    def plot_pareto(self, is_ground_truth=False):
        """Objective–variance Pareto front of all final solutions (minimize both).
        is_ground_truth=True uses BOTH the true f and the analytic σ²(x1,level) at the
        final design; False uses the observed Y_best_final and n_rep sample variance."""
        colors, cfgs = self._color_map()
        pts = np.array([[self._final_obj(r, is_ground_truth), self._final_var(r, is_ground_truth)]
                        for r in self.runs])
        keys = [self.cfg_key(r) for r in self.runs]
        fig, ax = plt.subplots(figsize=(7.5, 5.5))
        for cfg in cfgs:
            P = np.array([pts[i] for i in range(len(pts)) if keys[i] == cfg])
            if P.size == 0: continue
            st = self._style(cfg)
            ax.scatter(P[:, 0], P[:, 1], color=st["color"], marker=st["marker"], s=38, alpha=0.75,
                       edgecolor="k", linewidth=0.3, label=label(*cfg))
        order = np.argsort(pts[:, 0]); front, best_v = [], np.inf
        for i in order:
            if pts[i, 1] <= best_v:
                front.append(pts[i]); best_v = pts[i, 1]
        front = np.array(front)
        ax.plot(front[:, 0], front[:, 1], "k--", lw=1.3, label="Pareto front")
        ax.scatter(front[:, 0], front[:, 1], facecolors="none", edgecolors="k", s=110, lw=1.3)
        ax.set_xlabel("TRUE final objective $f(x^*)$" if is_ground_truth else "Final objective  $y$ (noisy)")
        ax.set_ylabel("TRUE noise variance σ²(x*)" if is_ground_truth
                      else "Final aleatoric variance (sample estimate)")
        ax.set_title(f"Objective–variance Pareto front of final solutions\n{_src_tag(is_ground_truth)}")
        ax.legend(fontsize=8, loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
        ax.grid(alpha=0.3); fig.tight_layout()
        return fig

    def plot_incumbent_variance(self, is_ground_truth=False):
        """Aleatoric variance of the best-so-far (incumbent) design vs iteration.
        is_ground_truth=True: rank the incumbent by the TRUE f AND plot the analytic
        TRUE noise variance σ²(x1,level) at it. is_ground_truth=False: rank by the noisy
        sample-mean AND plot the n_rep sample-variance estimate (Y_var_sampled)."""
        colors, cfgs = self._color_map()
        n_reps = self.n_reps
        fig, axes = plt.subplots(1, len(n_reps), figsize=(6*len(n_reps), 4.5),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            for i, cfg in enumerate(cfgs):
                curves = []
                for r in self.runs:
                    if self.cfg_key(r) != cfg or r["n_rep"] != nr: continue
                    n0 = r["n_initial"]
                    key = self._true_obj_of_samples(r) if is_ground_truth else r["Y_sampled"]
                    v = self._true_var_of_samples(r) if is_ground_truth else r["Y_var_sampled"]
                    niter = len(v) - n0
                    curves.append([v[int(np.nanargmin(key[:n0+k]))] for k in range(1, niter+1)])
                if not curves: continue
                L = min(len(c) for c in curves); C = np.array([c[:L] for c in curves])
                st = self._style(cfg); me = max(1, L//6)
                ax.plot(np.arange(1, L+1), C.mean(0), color=st["color"], ls=st["linestyle"],
                        marker=st["marker"], markevery=me, ms=4.5, lw=1.6, label=label(*cfg))
            ax.set_title(f"n_rep = {nr}"); ax.set_xlabel("Iteration"); ax.grid(alpha=0.3)
        axes[0][0].set_ylabel("TRUE noise variance σ² of incumbent" if is_ground_truth
                              else "aleatoric variance of incumbent (sample est.)")
        fig.suptitle(f"Robustness: noise at the current best design vs iteration\n{_sel_tag(is_ground_truth)}")
        self._facet_legend(fig, axes[0][0])
        return fig

    def incumbent_variance_table(self, is_ground_truth=True, tol=0.15, excel_path=None):
        """Per-config × n_rep summary of plot_incumbent_variance: the aleatoric noise
        σ² at the incumbent (best-so-far) design, summarised over the 30 seeds.

        NOTE the target is the OPTIMUM's noise σ²_opt (the true optimum sits in a noisy
        region), NOT zero; the quietest point in the whole domain is far lower. So a
        converged incumbent settles AT σ²_opt; a value well above it = stuck in a noisy
        off-optimum region (non-convergence).
          final  — mean incumbent σ² at the last iteration (where it settles),
          median — robust settle point (high mean + low median = a few tail seeds),
          peak   — mean over seeds of each seed's WORST incumbent σ² (mid-run excursion;
                   low peak = risk-averse, never parks the incumbent in noisy extremes),
          conv%  — % of seeds whose final incumbent σ² is within tol of σ²_opt (settled
                   at the true-optimum noise; stuck-high or quiet-wrong seeds don't count).
        is_ground_truth matches plot_incumbent_variance. Pass excel_path for a styled
        .xlsx. Returns the DataFrame."""
        import pandas as pd
        lv, x1 = true_opt_location()
        s2opt = float(sigma(x1, lv)**2)
        n_reps = self.n_reps
        rows = {}
        for a, p in CONFIG_ORDER:
            cfg = canon_cfg(a, p)
            fam = "risk-neutral" if a in ("lcb", "pi", "ei") else "risk-aware"
            rec = {("", "family"): fam}
            for nr in n_reps:
                curves = []
                for r in self.runs:
                    if self.cfg_key(r) != cfg or r["n_rep"] != nr: continue
                    n0 = r["n_initial"]
                    key = self._true_obj_of_samples(r) if is_ground_truth else r["Y_sampled"]
                    v = self._true_var_of_samples(r) if is_ground_truth else r["Y_var_sampled"]
                    niter = len(v) - n0
                    curves.append([v[int(np.nanargmin(key[:n0+k]))] for k in range(1, niter+1)])
                if not curves:
                    for m in ("final", "median", "peak", "conv%"):
                        rec[(f"n_rep={nr}", m)] = np.nan
                    continue
                L = min(len(c) for c in curves); C = np.array([c[:L] for c in curves])
                finals = C[:, -1]; peaks = C.max(1)
                rec[(f"n_rep={nr}", "final")]  = float(np.mean(finals))
                rec[(f"n_rep={nr}", "median")] = float(np.median(finals))
                rec[(f"n_rep={nr}", "peak")]   = float(np.mean(peaks))
                rec[(f"n_rep={nr}", "conv%")]  = round(100 * np.mean(
                    np.abs(finals - s2opt) <= tol * s2opt))
            rows[label(a, p)] = rec
        df = pd.DataFrame(rows).T
        df.columns = pd.MultiIndex.from_tuples(df.columns)
        df.index.name = f"config (σ²_opt={s2opt:.3f}, tol=±{tol:.0%})"
        if excel_path:
            src = "TRUE σ²(x) at TRUE-ranked incumbent" if is_ground_truth \
                  else "n_rep sample-variance estimate at noisy-ranked incumbent"
            note = (f"incumbent aleatoric variance σ², src={src}; target σ²_opt={s2opt:.4f} "
                    f"(the OPTIMUM's noise, not 0); final/median/peak/conv% over 30 seeds; "
                    f"conv% = % seeds with final σ² within ±{tol:.0%} of σ²_opt")
            self._write_table_excel(df, excel_path, "incumbent_variance",
                                    {"final", "median", "peak"}, {"conv%"}, note)
            print("wrote", excel_path)
        return df

    # ============================ best designs ============================
    @staticmethod
    def _final_true_obj(run, source="recommended"):
        xr = run["X_best_final"] if source == "best" else run["X_min_est"][-1]
        lv = int(round(xr[1]))
        return float(f_true(np.array([xr[0]]), VAR_FCTR[lv-1])[0]) if 1 <= lv <= len(VAR_FCTR) else np.nan

    def plot_final_boxplots(self, source="recommended", is_ground_truth=True):
        """Final-objective distribution across seeds per acquisition, faceted by
        n_rep, vs the true global min. is_ground_truth=False -> biased best NOISY
        sample-mean; is_ground_truth=True -> TRUE noise-free f at the final design."""
        gt = ground_truth_min(); colors, cfgs = self._color_map(); n_reps = self.n_reps
        fig, axes = plt.subplots(1, len(n_reps), figsize=(5.5*len(n_reps), 4.5),
                                 squeeze=False, sharey=True)
        for ax, nr in zip(axes[0], n_reps):
            data, labs, cols = [], [], []
            for cfg in cfgs:
                if is_ground_truth:
                    vals = [self._final_true_obj(r, source) for r in self.runs
                            if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                else:
                    vals = [r["Y_best_final"] for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if vals:
                    data.append(vals); labs.append(label(*cfg)); cols.append(colors[cfg])
            bp = ax.boxplot(data, patch_artist=True, showmeans=True)
            for patch, c in zip(bp["boxes"], cols):
                patch.set_facecolor(c); patch.set_alpha(0.5)
            ax.axhline(gt, ls="--", color="k", lw=1.2, label="true global min")
            ax.set_xticklabels(labs, rotation=40, ha="right", fontsize=8)
            ax.set_title(f"n_rep = {nr}"); ax.grid(alpha=0.3, axis="y")
        kind = f"true objective at {source} design" if is_ground_truth else "noisy best sample-mean"
        axes[0][0].set_ylabel(f"final objective\n({kind})")
        axes[0][-1].legend(fontsize=8)
        fig.suptitle(f"Final-objective distribution across seeds\n{_src_tag(is_ground_truth)}")
        fig.tight_layout()
        return fig

    def plot_best_designs(self, n_rep=10, expected_cfgs=CONFIG_ORDER, is_ground_truth=False):
        """Best design per seed vs the true optimum, one panel per acquisition.
        is_ground_truth=False: the design BO reported (min noisy sample-mean).
        is_ground_truth=True:  the truly-best sampled design (min true f).
        expected_cfgs (default = the full 12-config grid) keeps every panel present."""
        colors, _ = self._color_map()
        all_cfgs = ([canon_cfg(a, p) for (a, p) in expected_cfgs]
                    if expected_cfgs is not None else self._color_map()[1])
        gt_lv, gt_x1 = true_opt_location(); nlev = len(VAR_FCTR)
        cols = min(4, len(all_cfgs)); rows = int(np.ceil(len(all_cfgs)/cols))
        fig, axes = plt.subplots(rows, cols, figsize=(3.8*cols, 3.4*rows),
                                 squeeze=False, sharex=True, sharey=True)
        rng = np.random.default_rng(0)
        for ax in axes.ravel():
            ax.set_xticks(range(1, nlev+1))
            ax.set_xticklabels([f"{i+1}\n({v:g})" for i, v in enumerate(VAR_FCTR)], fontsize=8)
            ax.set_xlim(0.5, nlev+0.5); ax.grid(axis="y", alpha=0.4, linestyle="--")
            ax.grid(axis="x", alpha=0)
            ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        for i, cfg in enumerate(all_cfgs):
            ax = axes.ravel()[i]
            sel = [r for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == n_rep]
            if is_ground_truth:
                B = np.array([r["X_sampled"][self._best_sample_idx(r, True)] for r in sel])
            else:
                B = np.array([r["X_best_final"] for r in sel])
            if len(B) > 0:
                lv = np.round(B[:, 1]).astype(int); jit = rng.uniform(-0.15, 0.15, size=len(lv))
                ax.scatter(lv + jit, B[:, 0], c=[colors.get(cfg, "gray")], s=40, alpha=0.5,
                           edgecolor="white", linewidth=0.5)
                ax.plot(gt_lv, gt_x1, marker="X", ms=5, c="black", alpha=0.7, zorder=5)
                frac = np.mean(lv == gt_lv) * 100
                ax.set_title(f"{label(*cfg)} — {frac:.0f}% optimal lvl", fontsize=10, weight="medium")
            else:
                ax.set_title(f"{label(*cfg)}", fontsize=10, weight="medium", color="gray")
                ax.text(0.5, 0.5, "Awaiting Data...", ha="center", va="center",
                        transform=ax.transAxes, color="gray", style="italic", fontsize=9)
        for i in range(len(all_cfgs), rows*cols):
            axes.ravel()[i].set_axis_off()
        for ax in axes[-1]:
            if ax.get_visible(): ax.set_xlabel("categorical level (value)", fontsize=9)
        for ax in axes[:, 0]: ax.set_ylabel("$x_1$", fontsize=10)
        legend_elements = [
            Line2D([0], [0], marker="o", color="w", markerfacecolor="gray", alpha=0.6,
                   markeredgecolor="white", markersize=10, label="Per-seed best"),
            Line2D([0], [0], marker="X", color="w", markerfacecolor="black", alpha=0.7,
                   markersize=7, label="True optimum")]
        fig.legend(handles=legend_elements, loc="lower center", ncol=2, frameon=False,
                   fontsize=11, bbox_to_anchor=(0.5, 0.01))
        fig.suptitle(f"Best design per seed, n_rep={n_rep}  ·  {_sel_tag(is_ground_truth)}",
                     fontsize=12, weight="bold", y=0.98)
        fig.tight_layout(rect=[0, 0.08, 1, 0.96])
        return fig

    def plot_best_per_category(self, n_rep=10, expected_cfgs=CONFIG_ORDER, is_ground_truth=False):
        """Best objective found in EACH category (mean ± 95% CI across seeds) vs the
        true per-category minimum, per acquisition (shows all 5 categories).
        is_ground_truth=False: lowest NOISY sample-mean per category.
        is_ground_truth=True:  best TRUE f among sampled designs per category."""
        base_colors, present = self._color_map()
        if expected_cfgs is not None:
            cfgs = [canon_cfg(a, p) for (a, p) in expected_cfgs]
            cmap = plt.get_cmap("tab20" if len(cfgs) > 10 else "tab10")
            colors = {c: cmap(i % cmap.N) for i, c in enumerate(cfgs)}
        else:
            cfgs = [c for c in present if any(self.cfg_key(r) == c and r["n_rep"] == n_rep for r in self.runs)]
            colors = base_colors
            if not cfgs: raise ValueError(f"no data at n_rep={n_rep}")
        tmin = true_min_per_category(); nlev = len(VAR_FCTR); x = np.arange(1, nlev+1)
        cols = min(3, len(cfgs)); rows = int(np.ceil(len(cfgs)/cols))
        fig, axes = plt.subplots(rows, cols, figsize=(4.2*cols, 3.1*rows),
                                 squeeze=False, sharex=True, sharey=True)
        for ax, cfg in zip(axes.ravel(), cfgs):
            ax.set_xticks(x)
            ax.set_xticklabels([f"{i+1}\n({v:g})" for i, v in enumerate(VAR_FCTR)], fontsize=7)
            ax.grid(alpha=0.3)
            per_cat = self.true_obj_per_category if is_ground_truth else self.best_obj_per_category
            M = np.array([per_cat(r) for r in self.runs
                          if self.cfg_key(r) == cfg and r["n_rep"] == n_rep])
            if M.size == 0:
                ax.set_title(f"{label(*cfg)}  (awaiting data)", fontsize=9, color="gray")
                ax.text(0.5, 0.5, "Awaiting Data...", ha="center", va="center",
                        transform=ax.transAxes, color="gray", style="italic", fontsize=8)
                continue
            cnt = np.sum(~np.isnan(M), 0)
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                mean = np.nanmean(M, 0)
                sem = np.nan_to_num(np.nanstd(M, 0, ddof=1) / np.sqrt(np.maximum(cnt, 1)))
            ax.plot(x, tmin, "kX--", ms=10, lw=1, label="true min / category", zorder=2)
            ax.errorbar(x, mean, yerr=1.96*sem, marker="o", ms=6, color=colors[cfg],
                        capsize=3, lw=1.5, label="best found (mean ± 95% CI)", zorder=3)
            ax.set_title(f"{label(*cfg)}  (n={M.shape[0]} seeds)", fontsize=9)
            ax.legend(fontsize=6.5, loc="upper left")
        for ax in axes[-1]: ax.set_xlabel("categorical level (value)")
        for ax in axes[:, 0]: ax.set_ylabel("best objective in category")
        for ax in axes.ravel()[len(cfgs):]: ax.set_visible(False)
        fig.suptitle(f"Best objective found within each category, n_rep={n_rep}\n{_src_tag(is_ground_truth)}")
        fig.tight_layout()
        return fig

    # ============================ grid / budget / cost ============================
    def plot_nrep_sensitivity(self, gt=None, is_ground_truth=False):
        """Final objective vs n_rep (mean ± 95% CI), one line per acquisition.
        is_ground_truth=True uses the TRUE noise-free f at the final design instead of Y_best_final."""
        if gt is None: gt = ground_truth_min()
        colors, cfgs = self._color_map(); n_reps = self.n_reps
        fig, ax = plt.subplots(figsize=(8.5, 5.5))
        for cfg in cfgs:
            xs, ys, es = [], [], []
            for nr in n_reps:
                vals = [self._final_obj(r, is_ground_truth) for r in self.runs
                        if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if vals:
                    xs.append(nr); ys.append(np.mean(vals))
                    es.append(1.96*np.std(vals, ddof=1)/np.sqrt(len(vals)) if len(vals) > 1 else 0)
            st = self._style(cfg)
            ax.errorbar(xs, ys, yerr=es, marker=st["marker"], ls=st["linestyle"], color=st["color"],
                        capsize=3, ms=6, lw=1.6, label=label(*cfg))
        ax.axhline(gt, ls=(0, (6, 4)), color="0.3", lw=1.2, label="ground truth")
        ax.set_xticks(n_reps); ax.set_xlabel("n_rep (replicates per location)")
        ax.set_ylabel("Final objective (mean ± 95% CI)")
        ax.set_title(f"Sensitivity to experimental budget (n_rep)\n{_src_tag(is_ground_truth)}")
        ax.legend(fontsize=8, loc="center left", bbox_to_anchor=(1.01, 0.5), frameon=False)
        ax.grid(alpha=0.3); fig.tight_layout()
        return fig

    # def plot_summary_heatmaps(self, is_ground_truth=False, log=False,
    #                           obj_cmap="viridis_r", var_cmap="magma_r"):
    #     """Grid summary heatmaps: mean final objective & mean final variance
    #     (rows = acquisition, cols = n_rep). is_ground_truth=True uses the TRUE f at the
    #     final design. log=True uses a logarithmic colour scale (applied per panel only
    #     where all values are positive — useful for the wide-ranging variance panel).
    #     obj_cmap / var_cmap pick the colormap for each panel."""
    #     from matplotlib.colors import LogNorm
    #     colors, cfgs = self._color_map(); n_reps = self.n_reps
    #     labs = [label(*c) for c in cfgs]
    #     Mobj = np.full((len(cfgs), len(n_reps)), np.nan)
    #     Mvar = np.full((len(cfgs), len(n_reps)), np.nan)
    #     for i, cfg in enumerate(cfgs):
    #         for j, nr in enumerate(n_reps):
    #             o = [self._final_obj(r, is_ground_truth) for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
    #             v = [r["Y_var_best_final"] for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
    #             if o: Mobj[i, j] = np.mean(o)
    #             if v: Mvar[i, j] = np.mean(v)
    #     otitle = "mean TRUE final objective" if is_ground_truth else "mean final objective (noisy)"
    #     fig, axes = plt.subplots(1, 2, figsize=(11, 0.5*len(cfgs)+3))
    #     for ax, M, title, cmap in [(axes[0], Mobj, otitle, obj_cmap),
    #                                 (axes[1], Mvar, "mean final variance", var_cmap)]:
    #         vals = M[np.isfinite(M)]
    #         norm = LogNorm(vmin=vals.min(), vmax=vals.max()) if (log and vals.size and vals.min() > 0) else None
    #         im = ax.imshow(M, cmap=cmap, aspect="auto", norm=norm)
    #         ax.set_xticks(range(len(n_reps))); ax.set_xticklabels([f"n_rep={n}" for n in n_reps])
    #         ax.set_yticks(range(len(cfgs))); ax.set_yticklabels(labs, fontsize=8)
    #         for i in range(M.shape[0]):
    #             for j in range(M.shape[1]):
    #                 if not np.isnan(M[i, j]):
    #                     ax.text(j, i, f"{M[i,j]:.2g}", ha="center", va="center", fontsize=7, color="w")
    #         ax.set_title(title + ("  (log scale)" if norm is not None else "")); fig.colorbar(im, ax=ax, shrink=0.8)
    #     scale = " — log colour scale" if log else ""
    #     fig.suptitle(f"Grid summary (rows = acquisition, cols = n_rep){scale}\n{_src_tag(is_ground_truth)}")
    #     return fig

    def plot_summary_heatmaps(self, is_ground_truth=False, log=False,
                          obj_cmap="viridis_r", var_cmap="magma_r",
                          obj_vmin=None, obj_vmax=None,   # Retained from previous update
                          var_vmin=None, var_vmax=None,
                          txtwidth=1, fontsize=15,
                          wspace=0.3
                          ):  # Retained from previous update
        """Grid summary heatmaps: mean final objective & mean final variance
        (rows = acquisition, cols = n_rep). is_ground_truth=True uses the TRUE f at the
        final design. log=True uses a logarithmic colour scale (applied per panel only
        where all values are positive — useful for the wide-ranging variance panel).
        obj_cmap / var_cmap pick the colormap for each panel.
        *_vmin / *_vmax allow manual colorbar limits."""
        from matplotlib.colors import LogNorm
        import matplotlib.patheffects as pe # <-- NEW: Import patheffects for text contour

        colors, cfgs = self._color_map(); n_reps = self.n_reps
        labs = [label(*c) for c in cfgs]
        Mobj = np.full((len(cfgs), len(n_reps)), np.nan)
        Mvar = np.full((len(cfgs), len(n_reps)), np.nan)
        
        for i, cfg in enumerate(cfgs):
            for j, nr in enumerate(n_reps):
                o = [self._final_obj(r, is_ground_truth) for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                v = [self._final_var(r, is_ground_truth) for r in self.runs if self.cfg_key(r) == cfg and r["n_rep"] == nr]
                if o: Mobj[i, j] = np.mean(o)
                if v: Mvar[i, j] = np.mean(v)

        otitle = "mean TRUE final objective" if is_ground_truth else "mean final objective (noisy)"
        vtitle = "mean TRUE noise variance σ²" if is_ground_truth else "mean final variance (sample est.)"
        fig, axes = plt.subplots(1, 2, figsize=(11, 0.5*len(cfgs)+3))

        for ax, M, title, cmap, v_min, v_max in [
            (axes[0], Mobj, otitle, obj_cmap, obj_vmin, obj_vmax),
            (axes[1], Mvar, vtitle, var_cmap, var_vmin, var_vmax)
        ]:
            vals = M[np.isfinite(M)]
            
            actual_vmin = v_min if v_min is not None else (vals.min() if vals.size else None)
            actual_vmax = v_max if v_max is not None else (vals.max() if vals.size else None)

            if log and vals.size and vals.min() > 0:
                norm = LogNorm(vmin=actual_vmin, vmax=actual_vmax)
                im = ax.imshow(M, cmap=cmap, aspect="auto", norm=norm)
            else:
                im = ax.imshow(M, cmap=cmap, aspect="auto", vmin=actual_vmin, vmax=actual_vmax)
                norm = None 
                
            ax.set_xticks(range(len(n_reps))); ax.set_xticklabels([f"n_rep={n}" for n in n_reps])
            ax.set_yticks(range(len(cfgs))); ax.set_yticklabels(labs, fontsize=8)
            
            for i in range(M.shape[0]):
                for j in range(M.shape[1]):
                    if not np.isnan(M[i, j]):
                        # <-- NEW: Added path_effects to the text rendering
                        ax.text(j, i, f"{M[i,j]:.2g}", ha="center", va="center", 
                                fontsize=fontsize, color="w",
                                path_effects=[pe.withStroke(linewidth=txtwidth, foreground="black")])
                                
            ax.set_title(title + ("  (log scale)" if norm is not None else "")); fig.colorbar(im, ax=ax, shrink=0.8)
            
        scale = " — log colour scale" if log else ""
        fig.suptitle(f"Grid summary (rows = acquisition, cols = n_rep){scale}\n{_src_tag(is_ground_truth)}")
        fig.subplots_adjust(wspace=wspace) 
        return fig

    def plot_runtime(self):
        """Mean wall-clock runtime per acquisition."""
        colors, cfgs = self._color_map()
        fig, ax = plt.subplots(figsize=(8, 4.5)); labs, means, cols = [], [], []
        for cfg in cfgs:
            rt = [r["runtime"] for r in self.runs if self.cfg_key(r) == cfg]
            if rt:
                labs.append(label(*cfg)); means.append(np.mean(rt)); cols.append(colors[cfg])
        ax.bar(range(len(labs)), means, color=cols, edgecolor="k")
        ax.set_xticks(range(len(labs))); ax.set_xticklabels(labs, rotation=40, ha="right", fontsize=8)
        ax.set_ylabel("mean runtime per run (s)"); ax.grid(alpha=0.3, axis="y")
        ax.set_title("Computational cost per acquisition"); fig.tight_layout()
        return fig

    # ============================ single run ============================
    def plot_single_run(self, mat_path, cmap="viridis"):
        """Per-category slices of ONE run: true f + noise band + where BO sampled.
        cmap = the sequential colormap used for the BO-iteration colouring."""
        if not os.path.isabs(mat_path) and not os.path.exists(mat_path):
            mat_path = os.path.join(self.results_dir, mat_path)
        r = _load_run_file(mat_path)                  # handles .npz (this study) and .mat (LVGP)
        X = np.atleast_2d(r["X_sampled"]).astype(float); Y = np.ravel(r["Y_sampled"]).astype(float)
        n0 = r["n_initial"]; Xme = np.atleast_2d(r["X_min_est"]).astype(float)
        acf = r["acf"]; param = r["param"]; nrep = r["n_rep"]; seed = r["seed"]
        gt_lv, gt_x1 = true_opt_location(); x1g = np.linspace(LB, UB, 400)
        levels = np.round(X[:, 1]).astype(int); order = np.arange(len(X)); nlev = len(VAR_FCTR)
        fig, axes = plt.subplots(1, nlev, figsize=(3.6*nlev, 4.2), sharey=True)
        for lv in range(1, nlev+1):
            ax = axes[lv-1]; ft = f_true(x1g, VAR_FCTR[lv-1]); s = sigma(x1g, lv)
            ax.fill_between(x1g, ft-1.96*s, ft+1.96*s, color="0.85", label="true ±1.96σ")
            ax.plot(x1g, ft, "k-", lw=1.5, label="true f")
            sel = levels == lv; init = sel & (order < n0); bo = sel & (order >= n0)
            ax.scatter(X[init, 0], Y[init], c="0.5", marker="x", s=45, label="initial DOE")
            if bo.any():
                it = order[bo] - n0 + 1
                ax.scatter(X[bo, 0], Y[bo], c=it, cmap=cmap, s=42,
                           edgecolor="k", linewidth=0.4, zorder=3, label="BO samples")
            if int(round(Xme[-1, 1])) == lv:
                ax.axvline(Xme[-1, 0], color="r", ls=":", lw=1.6, label="recommended x*")
            if lv == gt_lv:
                ax.plot(gt_x1, f_true(np.array([gt_x1]), VAR_FCTR[lv-1])[0], "m*", ms=15,
                        zorder=4, label="global optimum")
            ax.set_title(f"level {lv} (val {VAR_FCTR[lv-1]:g}, noise×{NOISE_MULS[lv-1]:g})", fontsize=9)
            ax.set_xlabel("$x_1$"); ax.grid(alpha=0.3)
        axes[0].set_ylabel("objective $y$")
        plabel = "" if param != param else f"={param:g}"
        fig.suptitle(f"Single BO run — {acf}{plabel}, n_rep={nrep}, seed={seed}", y=1.02)
        fig.tight_layout()
        return fig


# ==================== dual-format loader + cross-study comparison ====================
def _load_run_file(f):
    """Load ONE run from .npz (this per-category-GP study) or .mat (study_v2 LVGP) into the
    normalized run dict used throughout StudyResults. Returns None if it has no results.
    The two studies share study_driver.m's field names, so only meta access differs."""
    if f.endswith(".npz"):
        src = np.load(f, allow_pickle=True)
        if "Y_min_history" not in src:
            return None
        meta = src["meta"].item()                     # plain dict
        get = lambda k: meta[k]
    else:
        src = scipy.io.loadmat(f)
        if "Y_min_history" not in src:
            return None
        meta = src["meta"][0, 0]                       # MATLAB struct
        get = lambda k: _s(meta[k])
    return dict(
        acf=str(get("acf")), param=float(get("acf_param")),
        n_rep=int(get("n_rep")), seed=int(get("seed")), runtime=float(get("runtime")),
        Y_min_history=np.ravel(src["Y_min_history"]).astype(float),
        X_sampled=np.atleast_2d(src["X_sampled"]).astype(float),
        n_initial=int(_s(src["n_initial"])),
        X_min_est=np.atleast_2d(src["X_min_est"]).astype(float),
        Y_sampled=np.ravel(src["Y_sampled"]).astype(float),
        Y_var_sampled=np.ravel(src["Y_var_sampled"]).astype(float),
        X_best_final=np.ravel(src["X_best_final"]).astype(float),
        Y_best_final=float(_s(src["Y_best_final"])),
        Y_var_best_final=float(_s(src["Y_var_best_final"])),
    )


def _cfg_param(cfg):
    """canon cfg key -> param value (nan for the 'na' baselines)."""
    return float("nan") if cfg[1] == "na" else float(cfg[1])


def _metric_trajectories(study, metric, n_rep, as_regret=True):
    """{cfg_key: (iters, mean, sem)} of the chosen metric across seeds at the given n_rep.
    Returns the best VALUE found so far, or its regret (value - f*) when as_regret=True.

    metric (or is_ground_truth in the callers) resolves to:
      'true'|'true_best_sampled' -- best TRUE objective among ALL sampled points,
            min_i f_true(X_sampled[i]) (>= f*); the noiseless curve.
      'noisy'|'best_y' -- best NOISY sample-mean Y_min_history (can be < f*).
      'true_regret' -- true f at the model's RECOMMENDED optimum X_min_est.
    """
    metric = {"true": "true_best_sampled", "noiseless": "true_best_sampled",
              "noisy": "best_y"}.get(metric, metric)
    from collections import defaultdict
    gt = ground_truth_min()
    series = defaultdict(list)
    for r in study.runs:
        if r["n_rep"] != n_rep:
            continue
        if metric == "true_best_sampled":
            Xs = r["X_sampled"]; n0 = int(r["n_initial"])
            ft = np.array([f_true(Xs[i, 0], VAR_FCTR[int(round(Xs[i, 1])) - 1])
                           for i in range(len(Xs))])
            cummin = np.minimum.accumulate(ft)                 # best TRUE value found so far
            niter = len(r["Y_min_history"])
            idx = np.clip(n0 - 1 + np.arange(1, niter + 1), 0, len(cummin) - 1)
            traj = cummin[idx]                                 # best TRUE value (>= f*)
        elif metric == "true_regret":
            Xme = r["X_min_est"]
            traj = np.array([f_true(Xme[i, 0], VAR_FCTR[int(round(Xme[i, 1])) - 1])
                             for i in range(Xme.shape[0])])     # true value at recommended opt
        elif metric == "best_y":
            traj = r["Y_min_history"]                           # best NOISY value
        else:
            raise ValueError(f"unknown metric {metric!r}")
        if as_regret:
            traj = traj - gt                                   # difference from f* (regret)
        series[StudyResults.cfg_key(r)].append(traj)
    out = {}
    for cfg, lst in series.items():
        L = min(len(t) for t in lst)
        A = np.array([t[:L] for t in lst])
        n = A.shape[0]
        sem = A.std(0, ddof=1) / np.sqrt(n) if n > 1 else np.zeros(L)
        out[cfg] = (np.arange(1, L + 1), A.mean(0), sem)
    return out


def compare_studies(study_a, study_b, metric="true_best_sampled", n_rep=10, configs=None,
                    labels=("Per-category GP", "LVGP"), ncols=3, logy=None, is_ground_truth=None,
                    as_regret=True):
    """Head-to-head overlay of two studies: one subplot per acquisition config, both studies'
    mean ± s.e. trajectory. Two booleans:
      is_ground_truth: True -> noiseless (best true f among samples), False -> noisy sample-mean.
      as_regret:       True -> plot value - f* (the difference; log axis when noiseless),
                       False -> plot the raw best value (converges to f*; linear, f* reference line).
    `configs` optionally restricts to acf_tags."""
    if is_ground_truth is not None:
        metric = "true_best_sampled" if is_ground_truth else "best_y"
    noiseless = metric in ("true_best_sampled", "true_regret")
    ta = _metric_trajectories(study_a, metric, n_rep, as_regret=as_regret)
    tb = _metric_trajectories(study_b, metric, n_rep, as_regret=as_regret)
    have = set(ta) & set(tb)
    cfgs = [canon_cfg(a, p) for a, p in CONFIG_ORDER]
    cfgs = [c for c in cfgs if c in have
            and (configs is None or acf_tag(c[0], _cfg_param(c)) in configs)]
    if not cfgs:
        raise ValueError("no shared (config, n_rep) cells between the two studies")
    logy = (as_regret and noiseless) if logy is None else logy   # only regret of true f is log-friendly
    kind = "true" if noiseless else "noisy"
    ylab = f"regret ({kind}) = value − f*" if as_regret else f"best {kind} value"
    nrows = int(np.ceil(len(cfgs) / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(5.2*ncols, 3.6*nrows), squeeze=False)
    for k, cfg in enumerate(cfgs):
        ax = axes[k // ncols][k % ncols]
        for (it, mean, sem), col, lab in [(ta[cfg], "C0", labels[0]), (tb[cfg], "C3", labels[1])]:
            m  = np.maximum(mean, 1e-12) if logy else mean
            lo = np.maximum(mean - sem, 1e-12) if logy else mean - sem
            ax.plot(it, m, color=col, lw=1.8, label=lab)
            ax.fill_between(it, lo, mean + sem, color=col, alpha=0.15)
        if not as_regret:
            ax.axhline(ground_truth_min(), color="0.4", ls=":", lw=1.0)   # f* reference
        ax.set_title(label(cfg[0], _cfg_param(cfg)))
        if logy:
            ax.set_yscale("log")
        ax.grid(alpha=0.25, which="both")
        if k % ncols == 0:
            ax.set_ylabel(ylab)
        if k // ncols == nrows - 1:
            ax.set_xlabel("BO iteration")
    for k in range(len(cfgs), nrows * ncols):
        axes[k // ncols][k % ncols].axis("off")
    axes[0][0].legend(fontsize=9)
    fig.suptitle(f"{labels[0]} vs {labels[1]} — {ylab} (n_rep={n_rep}, mean ± s.e. over seeds)", y=1.0)
    fig.tight_layout()
    return fig


def compare_studies_multi(studies_labels, metric="true_best_sampled", n_rep=10, configs=None,
                          ncols=3, logy=None, colors=("C0", "C2", "C3", "C1", "C4"),
                          is_ground_truth=None, as_regret=True):
    """Overlay N studies (one line each) per acquisition config -- the 3-way comparison.
    `is_ground_truth`: True -> noiseless, False -> noisy. `as_regret`: True -> value - f*
    (difference; log axis when noiseless), False -> raw best value (converges to f*; linear)."""
    if is_ground_truth is not None:
        metric = "true_best_sampled" if is_ground_truth else "best_y"
    noiseless = metric in ("true_best_sampled", "true_regret")
    trajs = [(_metric_trajectories(s, metric, n_rep, as_regret=as_regret), lab)
             for s, lab in studies_labels]
    have = set.intersection(*[set(td) for td, _ in trajs]) if trajs else set()
    cfgs = [canon_cfg(a, p) for a, p in CONFIG_ORDER]
    cfgs = [c for c in cfgs if c in have
            and (configs is None or acf_tag(c[0], _cfg_param(c)) in configs)]
    if not cfgs:
        raise ValueError("no (config, n_rep) cell shared by ALL studies")
    logy = (as_regret and noiseless) if logy is None else logy
    kind = "true" if noiseless else "noisy"
    ylab = f"regret ({kind}) = value − f*" if as_regret else f"best {kind} value"
    nrows = int(np.ceil(len(cfgs) / ncols))
    fig, axes = plt.subplots(nrows, ncols, figsize=(5.2*ncols, 3.6*nrows), squeeze=False)
    for k, cfg in enumerate(cfgs):
        ax = axes[k // ncols][k % ncols]
        for (td, lab), col in zip(trajs, colors):
            it, mean, sem = td[cfg]
            m  = np.maximum(mean, 1e-12) if logy else mean
            lo = np.maximum(mean - sem, 1e-12) if logy else mean - sem
            ax.plot(it, m, color=col, lw=1.8, label=lab)
            ax.fill_between(it, lo, mean + sem, color=col, alpha=0.13)
        if not as_regret:
            ax.axhline(ground_truth_min(), color="0.4", ls=":", lw=1.0)
        ax.set_title(label(cfg[0], _cfg_param(cfg)))
        if logy:
            ax.set_yscale("log")
        ax.grid(alpha=0.25, which="both")
        if k % ncols == 0:
            ax.set_ylabel(ylab)
        if k // ncols == nrows - 1:
            ax.set_xlabel("BO iteration")
    for k in range(len(cfgs), nrows * ncols):
        axes[k // ncols][k % ncols].axis("off")
    axes[0][0].legend(fontsize=9)
    fig.suptitle(f"{'  vs  '.join(l for _, l in studies_labels)} — {ylab} "
                 f"(n_rep={n_rep}, mean ± s.e. over seeds)", y=1.0)
    fig.tight_layout()
    return fig


def compare_summary_heatmaps(studies_labels, is_ground_truth=True, log=False,
                             obj_cmap="viridis_r", var_cmap="magma_r", fontsize=8):
    """Summary heatmaps (rows = acquisition, cols = n_rep) for N studies in ONE figure.
    Top row of panels = mean final OBJECTIVE per study; bottom row = mean noise VARIANCE σ².
    All objective panels share ONE colour range; all variance panels share another -- so the
    studies are directly comparable cell-by-cell. studies_labels = list of (StudyResults, label).
    log=True puts the variance row on a log colour scale."""
    from matplotlib.colors import LogNorm
    import matplotlib.patheffects as pe
    studies = [s for s, _ in studies_labels]
    labs = [l for _, l in studies_labels]
    n = len(studies)
    cfg_keys = [canon_cfg(a, p) for a, p in CONFIG_ORDER]
    cfg_labels = [label(a, p) for a, p in CONFIG_ORDER]
    n_reps = sorted(studies[0].n_reps)

    def grids(s):
        Mo = np.full((len(cfg_keys), len(n_reps)), np.nan)
        Mv = np.full((len(cfg_keys), len(n_reps)), np.nan)
        for i, cfg in enumerate(cfg_keys):
            for j, nr in enumerate(n_reps):
                o = [s._final_obj(r, is_ground_truth) for r in s.runs
                     if s.cfg_key(r) == cfg and r["n_rep"] == nr]
                v = [s._final_var(r, is_ground_truth) for r in s.runs
                     if s.cfg_key(r) == cfg and r["n_rep"] == nr]
                if o: Mo[i, j] = np.mean(o)
                if v: Mv[i, j] = np.mean(v)
        return Mo, Mv

    G = [grids(s) for s in studies]
    Mobjs = [g[0] for g in G]
    Mvars = [g[1] for g in G]

    def rng(Ms):
        vals = np.concatenate([M[np.isfinite(M)] for M in Ms]) if Ms else np.array([])
        return (float(vals.min()), float(vals.max())) if vals.size else (None, None)
    ovmin, ovmax = rng(Mobjs)            # SHARED objective range across studies
    vvmin, vvmax = rng(Mvars)            # SHARED variance range across studies

    fig, axes = plt.subplots(2, n, figsize=(3.3 * n + 1.0, 0.42 * len(cfg_keys) + 2.5),
                             squeeze=False, constrained_layout=True)
    row_im = [None, None]
    for col in range(n):
        for row, (M, cmap, vmin, vmax) in enumerate(
                [(Mobjs[col], obj_cmap, ovmin, ovmax), (Mvars[col], var_cmap, vvmin, vvmax)]):
            ax = axes[row][col]
            if log and row == 1 and vmin and vmin > 0:
                im = ax.imshow(M, cmap=cmap, aspect="auto", norm=LogNorm(vmin=vmin, vmax=vmax))
            else:
                im = ax.imshow(M, cmap=cmap, aspect="auto", vmin=vmin, vmax=vmax)
            row_im[row] = im
            ax.set_xticks(range(len(n_reps)))
            ax.set_xticklabels([f"nrep={k}" for k in n_reps], fontsize=7)
            if col == 0:
                ax.set_yticks(range(len(cfg_keys))); ax.set_yticklabels(cfg_labels, fontsize=7)
            else:
                ax.set_yticks([])
            for i in range(M.shape[0]):
                for j in range(M.shape[1]):
                    if np.isfinite(M[i, j]):
                        ax.text(j, i, f"{M[i, j]:.2g}", ha="center", va="center",
                                fontsize=fontsize, color="w",
                                path_effects=[pe.withStroke(linewidth=1, foreground="black")])
            if row == 0:
                ax.set_title(labs[col], fontsize=11)
    pre = "TRUE " if is_ground_truth else ""
    fig.colorbar(row_im[0], ax=list(axes[0]), shrink=0.85, location="right",
                 label=f"mean {pre}final objective")
    fig.colorbar(row_im[1], ax=list(axes[1]), shrink=0.85, location="right",
                 label=f"mean {pre}noise variance σ²" + ("  (log)" if log else ""))
    fig.suptitle("Summary heatmaps across studies — objective panels share one range, "
                 f"variance panels share another\n{_src_tag(is_ground_truth)}")
    return fig


def compare_runtime(studies_labels, logy=True, colors=("C0", "C2", "C3", "C1")):
    """Mean wall-clock runtime per BO run across studies: grouped bars, x = acquisition,
    one bar per study. Log y by default (LVGP/MATLAB and the Python GP studies differ ~10x).
    The legend shows each study's overall mean runtime per run. NOTE: LVGP is MATLAB while the
    GP studies are Python, so absolute magnitudes are NOT a same-engine comparison."""
    cfg_keys = [canon_cfg(a, p) for a, p in CONFIG_ORDER]
    cfg_labels = [label(a, p) for a, p in CONFIG_ORDER]
    x = np.arange(len(cfg_keys))
    nstud = len(studies_labels)
    w = 0.8 / max(nstud, 1)
    fig, ax = plt.subplots(figsize=(13, 5))
    for si, (s, lab) in enumerate(studies_labels):
        means = [np.mean([r["runtime"] for r in s.runs if s.cfg_key(r) == cfg]) or np.nan
                 if any(s.cfg_key(r) == cfg for r in s.runs) else np.nan for cfg in cfg_keys]
        ov = float(np.mean([r["runtime"] for r in s.runs])) if s.runs else float("nan")
        ax.bar(x + (si - (nstud - 1) / 2) * w, means, w, color=colors[si % len(colors)],
               edgecolor="k", lw=0.3, label=f"{lab}  (overall {ov:.0f} s/run)")
    if logy:
        ax.set_yscale("log")
    ax.set_xticks(x); ax.set_xticklabels(cfg_labels, rotation=40, ha="right", fontsize=8)
    ax.set_ylabel("mean runtime per BO run (s)" + ("  [log]" if logy else ""))
    ax.set_title("Wall-clock runtime per BO run across studies\n"
                 "(LVGP = MATLAB; per-category & categorical GP = Python — different engines)")
    ax.legend(fontsize=9); ax.grid(axis="y", alpha=0.3, which="both")
    fig.tight_layout()
    return fig


def runtime_summary(studies_labels):
    """Print conclusion-friendly wall-clock runtime stats across studies and return a per-study
    summary DataFrame (runs, mean/median/std/min/max s per run, total minutes & hours). Also
    prints the mean-runtime ratio vs the fastest study and the per-n_rep breakdown."""
    import pandas as pd
    rows = []
    for s, lab in studies_labels:
        rt = np.array([r["runtime"] for r in s.runs], float)
        if rt.size == 0:
            continue
        rows.append(dict(study=lab, runs=len(rt), mean_s=rt.mean(), median_s=float(np.median(rt)),
                         std_s=rt.std(), min_s=rt.min(), max_s=rt.max(),
                         total_min=rt.sum() / 60, total_hr=rt.sum() / 3600))
    df = pd.DataFrame(rows).set_index("study")
    if df.empty:
        print("no runs loaded"); return df
    fastest = df["mean_s"].idxmin()
    print("Wall-clock runtime per BO run (seconds) and TOTAL compute per study "
          f"({df['runs'].iloc[0]} runs each):\n")
    print(df.round(1).to_string())
    print(f"\nMean runtime relative to the fastest study ({fastest}):")
    for lab in df.index:
        print(f"  {lab:18s} {df.loc[lab, 'mean_s'] / df['mean_s'].min():5.1f}x   "
              f"({df.loc[lab, 'mean_s']:.0f} s/run, total {df.loc[lab, 'total_hr']:.1f} h)")
    print("\nMean s/run by n_rep:")
    for s, lab in studies_labels:
        if not s.runs:
            continue
        by = {nr: float(np.mean([r["runtime"] for r in s.runs if r["n_rep"] == nr]))
              for nr in sorted(s.n_reps)}
        print(f"  {lab:18s} " + "  ".join(f"n_rep={nr}: {v:5.0f}s" for nr, v in by.items()))
    return df
