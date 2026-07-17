"""
results.py -- load and compare the (function x model) grid. Reads BOTH .npz (python engine) and
.mat (matlab engine) cells from results/<function>/<model>/<acq>/nrep<NN>/seed<NN>, indexed by
(function, model, acf, param, n_rep, seed). The headline metric is `true_best_sampled` (min f_true
over the sampled points), consistent with the study_v2* studies.
"""
import os
import glob
import numpy as np

from . import problems as P
from . import acquisitions
from .models import MODELS


def _unwrap_cell(a):
    """Peel MATLAB 1x1 cell/object nesting down to the underlying array."""
    a = np.asarray(a)
    while a.dtype == object and a.size == 1:
        a = np.asarray(a.reshape(-1)[0])
    return a


def _load_cell(path):
    hyper_z = None                                            # LVGP latent embedding (matlab cells only)
    if path.endswith(".npz"):
        m = np.load(path, allow_pickle=True)
        meta = m["meta"].item()
        g = lambda k: np.asarray(m[k])
    else:
        import scipy.io
        d = scipy.io.loadmat(path)
        mm = d["meta"][0, 0]
        meta = {}
        for k in mm.dtype.names:
            v = mm[k]
            meta[k] = str(v[0]) if v.dtype.kind in "US" else float(np.ravel(v)[0])
        g = lambda k: np.asarray(d[k])
        try:
            z = _unwrap_cell(d["hyper"][0, 0]["z"]).astype(float)
            hyper_z = z.reshape(-1, 2) if z.size else None    # (n_levels, dim_z=2)
        except Exception:
            hyper_z = None

    def opt(k, flat=True):
        """A schema-v2 field, or None on a v1 cell that predates it."""
        try:
            v = np.asarray(g(k), float)
        except Exception:
            return None
        return np.ravel(v) if flat else v

    out = dict(
        problem=str(meta.get("problem")), model=str(meta.get("model")), acf=str(meta.get("acf")),
        param=float(meta.get("acf_param", float("nan"))), n_rep=int(float(meta.get("n_rep"))),
        seed=int(float(meta.get("seed"))), runtime=float(meta.get("runtime", 0.0)),
        schema_version=int(float(meta.get("schema_version", 1))),
        Y_min_history=np.ravel(g("Y_min_history")).astype(float),
        X_sampled=np.atleast_2d(g("X_sampled").astype(float)),      # (n, d+1), level = LAST col
        Y_sampled=np.ravel(g("Y_sampled")).astype(float),      # replicate MEAN at each sampled point
        X_min_est=np.atleast_2d(g("X_min_est").astype(float)),
        Y_var_sampled=np.ravel(g("Y_var_sampled")).astype(float),
        n_initial=int(np.ravel(g("n_initial"))[0]),
    )
    out.update(                                               # ---- schema v2 (None on a v1 cell) ----
        Y_rep_sampled=opt("Y_rep_sampled", flat=False),       # (n_tr, n_rep) raw replicates
        acf_val=opt("acf_val"),                               # acquisition value at the chosen point
        mu_at_est=opt("mu_at_est"), s_at_est=opt("s_at_est"), r_at_est=opt("r_at_est"),
        f_true_sampled=opt("f_true_sampled"), sigma_true_sampled=opt("sigma_true_sampled"),
        X_init=opt("X_init", flat=False), Y_init=opt("Y_init"),
        Y_rep_init=opt("Y_rep_init", flat=False),
        hyper_z=hyper_z,                                       # LVGP latent embedding (None for python)
    )
    return out


class GridResults:
    def __init__(self, root, runs):
        self.root = root
        self.runs = runs

    @classmethod
    def load(cls, root="results"):
        runs = []
        for f in (glob.glob(os.path.join(root, "**", "*.npz"), recursive=True)
                  + glob.glob(os.path.join(root, "**", "*.mat"), recursive=True)):
            try:
                runs.append(_load_cell(f))
            except Exception as e:
                print(f"  skip {f}: {e}")
        return cls(root, runs)

    def functions(self):
        return sorted({r["problem"] for r in self.runs})

    def models(self):
        return sorted({r["model"] for r in self.runs})

    def select(self, function=None, model=None, acf=None, param=None, n_rep=None):
        out = self.runs
        if function is not None: out = [r for r in out if r["problem"] == function]
        if model is not None:    out = [r for r in out if r["model"] == model]
        if acf is not None:      out = [r for r in out if r["acf"] == acf]
        if n_rep is not None:    out = [r for r in out if r["n_rep"] == n_rep]
        if param is not None and param == param:
            out = [r for r in out if abs(r["param"] - param) < 1e-9]
        return out


def _true_best_traj(run, spec):
    """Per-iteration true_best_sampled = cumulative min f_true over sampled points."""
    X = run["X_sampled"]
    ft = np.array([float(np.ravel(spec.f_true_level(x[:-1], int(round(x[-1])))) [0]) for x in X])
    cummin = np.minimum.accumulate(ft)
    n0 = run["n_initial"]
    niter = len(run["Y_min_history"])
    idx = np.clip(n0 - 1 + np.arange(1, niter + 1), 0, len(ft) - 1)
    return cummin[idx]


def compare_models_on_function(grid, function, acf="ei", param=float("nan"), n_rep=10,
                               as_regret=True, logy=True, ax=None):
    """Overlay each model's true-value convergence (mean +/- s.e. over seeds) for one (function,
    acquisition, n_rep). as_regret -> value - f*; logy -> log axis. Returns the axes."""
    import matplotlib.pyplot as plt
    spec = P.get(function)
    fstar = P.ground_truth_min(spec)
    if ax is None:
        _, ax = plt.subplots(figsize=(8, 5))
    colors = ["C0", "C3", "C2", "C1", "C4", "C5"]
    for m, col in zip(sorted(grid.models()), colors):
        runs = grid.select(function=function, model=m, acf=acf, param=param, n_rep=n_rep)
        if not runs:
            continue
        trajs = [_true_best_traj(r, spec) for r in runs]
        L = min(len(t) for t in trajs)
        A = np.array([t[:L] for t in trajs])
        A = (A - fstar) if as_regret else A
        mean = A.mean(0); sem = A.std(0, ddof=1) / np.sqrt(len(A)) if len(A) > 1 else np.zeros(L)
        lab = MODELS[m].label if m in MODELS else m
        x = np.arange(1, L + 1)
        m_ = np.maximum(mean, 1e-6) if (logy and as_regret) else mean
        ax.plot(x, m_, color=col, lw=2, label=f"{lab} (n={len(A)})")
        ax.fill_between(x, np.maximum(mean - sem, 1e-6) if (logy and as_regret) else mean - sem,
                        mean + sem, color=col, alpha=0.15)
    if not as_regret:
        ax.axhline(fstar, color="grey", ls="--", lw=1, label=f"f* = {fstar:.4f}")
    if logy and as_regret:
        ax.set_yscale("log")
    ax.set_xlabel("BO iteration")
    ax.set_ylabel("regret = value − f*" if as_regret else "best true value")
    ax.set_title(f"{function} — {acquisitions.label(acf, param)} (n_rep={n_rep})")
    ax.grid(alpha=0.3, which="both"); ax.legend(fontsize=8)
    return ax


def final_regret_table(grid, acf="ei", param=float("nan"), n_rep=10):
    """Mean +/- s.e. final true_best regret per (function, model) for one acquisition."""
    import pandas as pd
    rows = {}
    for fn in grid.functions():
        spec = P.get(fn); fstar = P.ground_truth_min(spec)
        row = {}
        for m in grid.models():
            runs = grid.select(function=fn, model=m, acf=acf, param=param, n_rep=n_rep)
            if not runs:
                row[MODELS[m].label if m in MODELS else m] = None; continue
            fr = np.array([_true_best_traj(r, spec)[-1] - fstar for r in runs])
            row[MODELS[m].label if m in MODELS else m] = f"{fr.mean():.3f} ± {fr.std():.3f}"
        rows[fn] = row
    return pd.DataFrame(rows).T


# ======================================================================================
#  Multi-function views (ports the previous_ver/compare_methods.ipynb plots to the grid).
#  Every function below takes the FULL grid and facets over `functions` -- so the same call
#  works for 1 or all 10 test problems.
#
#  GROUND_TRUTH toggle, as in the old notebook:
#    True  -> the incumbent is chosen by the NOISE-FREE f, and sigma^2 is the TRUE noise variance
#    False -> the incumbent is chosen by the observed replicate mean, and sigma^2 is its sample variance
#  The `_true_best_traj` above is the ground-truth trajectory; `_noisy_best_traj` is its counterpart.
# ======================================================================================
MODEL_COLORS = {"standard_LVGP": "C0", "heter_LVGP": "C3",
                "separate_gp": "C2", "categorical_kernel": "C1",
                "lvgp_native": "#17becf",        # Python homoscedastic (cyan, pairs with C0)
                "heter_lvgp_native": "#9467bd",   # Python heteroscedastic (purple, pairs with C3)
                "lvgp_torch": "#7f7f7f", "heter_lvgp_torch": "#8c564b"}
MODEL_ORDER = ["standard_LVGP", "heter_LVGP", "separate_gp", "categorical_kernel"]

# Per-problem deep dive grammar: ACQUISITION -> colour, MODEL -> (linestyle, marker). Kept consistent
# across every panel so a reader learns the encoding once. Acquisition colours follow CONFIG_ORDER.
ACQ_COLORS = {a: f"C{i}" for i, (a, _p) in enumerate(acquisitions.CONFIG_ORDER)}
MODEL_STYLES = {"standard_LVGP": ("-", "o"), "heter_LVGP": ("--", "s"),
                "separate_gp": (":", "^"), "categorical_kernel": ("-.", "D"),
                "lvgp_native": ("-", "v"), "heter_lvgp_native": ("--", "P"),
                "lvgp_torch": (":", "X"), "heter_lvgp_torch": ("-.", "*")}


def _mlabel(m):
    return MODELS[m].label if m in MODELS else m


def _ordered_models(grid):
    present = set(grid.models())
    return [m for m in MODEL_ORDER if m in present] + sorted(present - set(MODEL_ORDER))


def _iter_slice(run, arr):
    """Restrict a per-SAMPLED-POINT array to the BO iterations (drop the initial DOE block)."""
    n0, niter = run["n_initial"], len(run["Y_min_history"])
    idx = np.clip(n0 - 1 + np.arange(1, niter + 1), 0, len(arr) - 1)
    return arr[idx], idx


def _noisy_best_traj(run, spec):
    """Cumulative min of the OBSERVED replicate mean (what the optimizer actually sees)."""
    return _iter_slice(run, np.minimum.accumulate(run["Y_sampled"]))[0]


def _incumbent_idx(run, spec, ground_truth):
    """Index (into X_sampled) of the incumbent best design after each BO iteration."""
    X = run["X_sampled"]
    if ground_truth:
        vals = run["f_true_sampled"]
        if vals is None:                                   # v1 cell -> recompute from the spec
            vals = np.array([float(np.ravel(spec.f_true_level(x[:-1], int(round(x[-1]))))[0]) for x in X])
    else:
        vals = run["Y_sampled"]
    argmins = np.empty(len(vals), int)                     # running argmin (no float-equality tricks)
    cur = 0
    for i, v in enumerate(vals):
        if v < vals[cur]:
            cur = i
        argmins[i] = cur
    return _iter_slice(run, argmins)[0]


def sigma2_at_best_traj(run, spec, ground_truth=True):
    """Noise VARIANCE at the incumbent best design, per BO iteration.
    ground_truth -> the true sigma(x,level)^2; else the replicate sample variance actually observed."""
    idx = _incumbent_idx(run, spec, ground_truth)
    if ground_truth:
        sig = run["sigma_true_sampled"]
        if sig is None:
            X = run["X_sampled"]
            sig = np.array([float(np.ravel(spec.sigma_level(x[:-1], int(round(x[-1]))))[0]) for x in X])
        return sig[idx] ** 2
    return run["Y_var_sampled"][idx]


def _panel(ax, grid, function, acf, param, n_rep, series_fn, ylabel, logy, title_extra="",
           center="mean"):
    """Shared plumbing: overlay every model's trajectory on one axes.
    center='mean'   -> mean with a +/- standard-error band ("expected performance")
    center='median' -> MEDIAN with an inter-quartile (25-75%) band ("typical run").
    They disagree when the per-seed distribution is skewed -- which it often is here (a few bad seeds
    inflate the mean), so the median view is the honest one for heavy-tailed regret."""
    spec = P.get(function)
    any_data = False
    for m in _ordered_models(grid):
        runs = grid.select(function=function, model=m, acf=acf, param=param, n_rep=n_rep)
        if not runs:
            continue
        trajs = [series_fn(r, spec) for r in runs]
        L = min(len(t) for t in trajs)
        A = np.array([t[:L] for t in trajs])
        if center == "median":
            mid = np.median(A, 0)
            lo_b, hi_b = np.percentile(A, 25, axis=0), np.percentile(A, 75, axis=0)
        else:
            mid = A.mean(0)
            sem = A.std(0, ddof=1) / np.sqrt(len(A)) if len(A) > 1 else np.zeros(L)
            lo_b, hi_b = mid - sem, mid + sem
        x = np.arange(1, L + 1)
        lo = np.maximum(lo_b, 1e-12) if logy else lo_b
        ax.plot(x, np.maximum(mid, 1e-12) if logy else mid,
                color=MODEL_COLORS.get(m, "C7"), lw=2, label=f"{_mlabel(m)} (n={len(A)})")
        ax.fill_between(x, lo, hi_b, color=MODEL_COLORS.get(m, "C7"), alpha=0.15)
        any_data = True
    if logy:
        ax.set_yscale("log")
    if not any_data:
        ax.text(0.5, 0.5, "no data yet", ha="center", va="center", transform=ax.transAxes,
                color="crimson", fontsize=11)
    ax.set_xlabel("BO iteration"); ax.set_ylabel(ylabel)
    ax.set_title(f"{function}{title_extra}", fontsize=10)
    ax.grid(alpha=0.3, which="both")
    if any_data:
        ax.legend(fontsize=7)
    return any_data


def facet(grid, kind, functions=None, acf="ei", param=float("nan"), n_rep=10,
          ground_truth=True, logy=True, ncol=2, figsize=(6.4, 4.4), center="mean"):
    """One panel per test function, models overlaid. kind:
         'regret'   value - f*            (log axis when logy)
         'value'    best true value, with the f* horizontal line
         'sigma2'   noise variance at the incumbent best design
    """
    import matplotlib.pyplot as plt
    fns = functions or grid.functions()
    nrow = (len(fns) + ncol - 1) // ncol
    fig, axes = plt.subplots(nrow, ncol, figsize=(figsize[0] * ncol, figsize[1] * nrow), squeeze=False)
    tag = "true f" if ground_truth else "observed mean"
    for ax, fn in zip(axes.ravel(), fns):
        spec = P.get(fn)
        fstar = P.ground_truth_min(spec)
        if kind == "regret":
            base = _true_best_traj if ground_truth else _noisy_best_traj
            _panel(ax, grid, fn, acf, param, n_rep, lambda r, s: base(r, s) - fstar,
                   "regret = value − f*", logy, center=center)
        elif kind == "value":
            base = _true_best_traj if ground_truth else _noisy_best_traj
            ok = _panel(ax, grid, fn, acf, param, n_rep, base, "best value", False, center=center)
            if ok:
                ax.axhline(fstar, color="grey", ls="--", lw=1.2)
                ax.annotate(f"f* = {fstar:.3f}", xy=(0.98, fstar), xycoords=("axes fraction", "data"),
                            ha="right", va="bottom", fontsize=7, color="grey")
        elif kind == "sigma2":
            _panel(ax, grid, fn, acf, param, n_rep,
                   lambda r, s: sigma2_at_best_traj(r, s, ground_truth),
                   "σ² at incumbent", logy, center=center)
        else:
            raise ValueError(f"unknown kind {kind!r}")
    for k in range(len(fns), nrow * ncol):
        axes.ravel()[k].axis("off")
    fig.suptitle(f"{kind} — {acquisitions.label(acf, param)}, n_rep={n_rep}, incumbent by {tag}",
                 y=1.002, fontsize=12)
    fig.tight_layout()
    return fig


def heatmaps_by_function(grid, functions=None, n_rep=10, ground_truth=True, ncol=2,
                         configs=None, annot_fmt=None, metric="regret"):
    """One conventional heatmap per problem: rows = acquisition, cols = model, cell = final `metric`
    ('regret' = value − f*, or 'noise' = σ² at the incumbent; both lower = better). Sequential viridis
    colour scale with a real value colorbar per panel. Rows/configs with no data anywhere are dropped."""
    import matplotlib.pyplot as plt
    fns = functions or grid.functions()
    cfgs = configs or acquisitions.CONFIG_ORDER
    models = _ordered_models(grid)
    fmt = annot_fmt or ("{:.3f}" if metric == "regret" else "{:.3g}")
    lab = "regret" if metric == "regret" else "σ² at incumbent"
    nrow = (len(fns) + ncol - 1) // ncol
    fig, axes = plt.subplots(nrow, ncol, figsize=(1 * len(models) + 3.2, 2.9 * nrow), squeeze=False)
    for ax, fn in zip(axes.ravel(), fns):
        spec = P.get(fn); fstar = P.ground_truth_min(spec)
        rows, labels = [], []
        for (a, p) in cfgs:
            row = []
            for m in models:
                runs = (grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
                        if a in MODELS[m].supports else [])
                row.append(np.mean([_final_metric(r, spec, fstar, metric, ground_truth) for r in runs])
                           if runs else np.nan)
            if not np.all(np.isnan(row)):
                rows.append(row); labels.append(acquisitions.label(a, p))
        if not rows:
            ax.text(0.5, 0.5, "no data yet", ha="center", va="center", transform=ax.transAxes,
                    color="crimson"); ax.set_title(fn, fontsize=10); ax.axis("off"); continue
        M = np.array(rows)
        im = ax.imshow(np.ma.masked_invalid(M), cmap="viridis_r", aspect="auto")
        ax.set_facecolor("0.85")                              # NaN (unsupported acq × model) shows grey
        ax.set_xticks(range(len(models))); ax.set_xticklabels([_mlabel(m) for m in models],
                                                              rotation=30, ha="right", fontsize=7)
        ax.set_yticks(range(len(labels))); ax.set_yticklabels(labels, fontsize=7)
        mmean = np.nanmean(M)
        for i in range(M.shape[0]):
            for j in range(M.shape[1]):
                if not np.isnan(M[i, j]):
                    ax.text(j, i, fmt.format(M[i, j]), ha="center", va="center", fontsize=6,
                            color="w" if M[i, j] > mmean else "k")
        ax.set_title(f"{fn}  (final {lab}, n_rep={n_rep})", fontsize=9)
        fig.colorbar(im, ax=ax, fraction=0.046)
    for k in range(len(fns), nrow * ncol):
        axes.ravel()[k].axis("off")
    fig.tight_layout()
    return fig


def regret_heatmap(grid, functions=None, n_rep=10, ground_truth=True, configs=None,
                   normalize="row", annot_fmt="{:.3f}", figsize=None):
    """SINGLE consolidated heatmap of final regret. Rows = (problem x acquisition), columns = models;
    every cell is ANNOTATED with the mean final regret (value - f*) for that model/acq/problem.

    Because the problems differ by orders of magnitude in regret scale, a shared raw color scale would
    wash out. COLOR is therefore normalized per `normalize`:
      'row'  : within each (problem, acq) row, best model -> green (0), worst -> red (1). Answers
               "which model wins this row" at a glance; the printed number gives the magnitude.
      'func' : min-max within each PROBLEM (comparable colors across that problem's acquisitions).
      'none' : one raw scale (only sensible for a single problem).
    Cells with no data are left blank/grey. Returns the figure."""
    import matplotlib.pyplot as plt
    from matplotlib.colors import ListedColormap
    fns = functions or grid.functions()
    cfgs = configs or acquisitions.CONFIG_ORDER
    models = _ordered_models(grid)
    base = _true_best_traj if ground_truth else _noisy_best_traj

    raw, ylabels, fkey, boundaries = [], [], [], []      # boundaries: y-index where a new problem starts
    for fn in fns:
        spec = P.get(fn); fstar = P.ground_truth_min(spec)
        started = False
        for (a, p) in cfgs:
            row = []
            for m in models:
                runs = grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
                row.append(np.mean([base(r, spec)[-1] - fstar for r in runs]) if runs else np.nan)
            if np.all(np.isnan(row)):
                continue
            if not started:
                boundaries.append(len(raw)); started = True
            raw.append(row); fkey.append(fn)
            ylabels.append(f"{fn} · {acquisitions.label(a, p)}")
    if not raw:
        fig, ax = plt.subplots(figsize=(5, 2)); ax.text(0.5, 0.5, "no data yet", ha="center",
            va="center", color="crimson", transform=ax.transAxes); ax.axis("off"); return fig
    R = np.array(raw, float)                              # (rows, models) raw regret

    # colour matrix C in [0,1], best=0 -> green, worst=1 -> red, computed per the chosen scope
    C = np.full_like(R, np.nan)
    if normalize == "row":
        groups = [[i] for i in range(len(R))]
    elif normalize == "func":
        groups = [[i for i in range(len(R)) if fkey[i] == fn] for fn in dict.fromkeys(fkey)]
    else:
        groups = [list(range(len(R)))]
    for g in groups:
        block = R[g]; finite = block[np.isfinite(block)]
        lo, hi = (finite.min(), finite.max()) if finite.size else (0.0, 1.0)
        C[g] = 0.0 if hi <= lo else (block - lo) / (hi - lo)

    fig, ax = plt.subplots(figsize=figsize or (1.7 * len(models) + 3.0, 0.42 * len(R) + 1.6))
    im = ax.imshow(np.ma.masked_invalid(C), cmap="RdYlGn_r", aspect="auto", vmin=0, vmax=1)
    ax.set_facecolor("0.85")                              # NaN cells show as grey
    ax.set_xticks(range(len(models))); ax.set_xticklabels([_mlabel(m) for m in models],
                                                          rotation=25, ha="right", fontsize=8)
    ax.set_yticks(range(len(R))); ax.set_yticklabels(ylabels, fontsize=7)
    for b in boundaries[1:]:                              # separators between problems
        ax.axhline(b - 0.5, color="k", lw=1.3)
    for i in range(R.shape[0]):
        for j in range(R.shape[1]):
            if np.isfinite(R[i, j]):
                ax.text(j, i, annot_fmt.format(R[i, j]), ha="center", va="center", fontsize=7,
                        color="k" if 0.30 < C[i, j] < 0.85 else "w")
    ax.set_title(f"final regret (value − f*) per model / acq — n_rep={n_rep}, "
                 f"incumbent by {'true f' if ground_truth else 'observed mean'}\n"
                 f"colour = rank within each {'row' if normalize=='row' else normalize} "
                 f"(green = best model, red = worst); number = regret", fontsize=9)
    cb = fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02)
    cb.set_ticks([0, 1]); cb.set_ticklabels(["best", "worst"])
    fig.tight_layout()
    return fig


def coverage(grid, n_rep=None):
    """How many seeds exist per (function, model, acquisition) -- the sweep is long, so know what
    the plots above are actually averaging over."""
    import pandas as pd
    rows = []
    for r in grid.runs:
        if n_rep is not None and r["n_rep"] != n_rep:
            continue
        rows.append(dict(function=r["problem"], model=_mlabel(r["model"]),
                         acq=acquisitions.label(r["acf"], r["param"]), seed=r["seed"]))
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    return df.pivot_table(index=["function", "acq"], columns="model", values="seed",
                          aggfunc="count", fill_value=0)


def runtime_table(grid):
    """Mean wall-clock seconds per cell, function x model."""
    import pandas as pd
    rows = []
    for r in grid.runs:
        rows.append(dict(function=r["problem"], model=_mlabel(r["model"]), runtime=r["runtime"]))
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).pivot_table(index="function", columns="model",
                                          values="runtime", aggfunc="mean").round(1)


# ======================================================================================
#  PER-PROBLEM DEEP DIVE
#  A focused report for a single test problem: the landscape + initial DOE, per-model
#  convergence, the multi-acquisition overlay (acq=colour, model=style), the regret-vs-noise
#  trade-off, and the actual sampling trajectory. `problem_report` assembles + saves them;
#  `all_problem_reports` runs it for every problem into plots/<problem>/.
# ======================================================================================
def _one_run(grid, function, model, acf, param, n_rep, seed):
    for r in grid.runs:
        if (r["problem"] == function and r["model"] == model and r["acf"] == acf
                and r["n_rep"] == n_rep and r["seed"] == seed
                and (param != param or abs(r["param"] - param) < 1e-9)):
            return r
    return None


def _init_doe_run(grid, function, n_rep, seed):
    """Any cell for (function, n_rep, seed) carries the SHARED initial design (X_init/Y_rep_init)."""
    for r in grid.runs:
        if (r["problem"] == function and r["n_rep"] == n_rep and r["seed"] == seed
                and r.get("X_init") is not None):
            return r
    return None


def problem_landscape(function, grid=None, n_rep=10, seed=1, ax=None, show_doe=True):
    """The problem itself: noise-free f(x1|level) per category (coloured by level) with the true noise
    band f ± σ shaded, the global optimum starred, and -- if a matching cell exists -- the shared
    initial DOE (design points + their noisy replicates) overlaid so you see what every model started
    from."""
    import matplotlib.pyplot as plt
    spec = P.get(function)
    if ax is None:
        _, ax = plt.subplots(figsize=(7, 5))
    x = np.linspace(spec.lb, spec.ub, 400)
    for lv in spec.levels:
        c = plt.cm.tab10((lv - 1) % 10)
        f = spec.f_true_level(x, lv); sig = spec.sigma_level(x, lv)
        ax.plot(x, f, color=c, lw=1.6, label=f"level {lv}", zorder=3)
        ax.fill_between(x, f - sig, f + sig, color=c, alpha=0.10, zorder=1)   # true ±1σ noise band
    lv_opt, x_opt = P.true_opt_location(spec)
    ax.scatter([x_opt], [spec.f_true_level(x_opt, lv_opt)], marker="*", s=240, c="crimson",
               edgecolor="k", lw=0.6, zorder=6, label=f"opt (lv {lv_opt})")
    if show_doe and grid is not None:
        run = _init_doe_run(grid, function, n_rep, seed)
        if run is not None:
            Xi, Yr = run["X_init"], run["Y_rep_init"]
            for i in range(len(Xi)):
                c = plt.cm.tab10((int(Xi[i, 1]) - 1) % 10)
                ax.scatter(np.full(Yr.shape[1], Xi[i, 0]), Yr[i], s=10, color=c, alpha=0.35, zorder=4)
                ax.scatter([Xi[i, 0]], [Yr[i].mean()], s=55, color=c, edgecolor="k", lw=0.5, zorder=5)
    ax.set_xlabel("x1"); ax.set_ylabel("f")
    ax.set_title(f"{function}: landscape + initial DOE (seed {seed})", fontsize=10)
    ax.grid(alpha=0.3); ax.legend(fontsize=7, ncol=2)
    return ax


def noise_per_level(function, ax=None, as_variance=False, mark_opt=True):
    """The heteroscedastic noise structure: the TRUE noise std σ(x1│level) (or σ² with as_variance)
    for each categorical level over the domain, coloured to match the landscape. This is the quantity
    the noise-aware models try to learn -- where in (x1, level) space the objective is noisy."""
    import matplotlib.pyplot as plt
    spec = P.get(function)
    if ax is None:
        _, ax = plt.subplots(figsize=(7, 5))
    x = np.linspace(spec.lb, spec.ub, 400)
    for lv in spec.levels:
        c = plt.cm.tab10((lv - 1) % 10)
        s = spec.sigma_level(x, lv)
        ax.plot(x, s ** 2 if as_variance else s, color=c, lw=1.8, label=f"level {lv}", zorder=3)
    if mark_opt:
        lv_opt, x_opt = P.true_opt_location(spec)
        ax.axvline(x_opt, color="crimson", ls="--", lw=1, zorder=2,
                   label=f"optimum x1 (lv {lv_opt})")
    ax.set_xlabel("x1"); ax.set_ylabel("σ² true noise" if as_variance else "σ true noise std")
    ax.set_title(f"{function}: heteroscedastic noise per level", fontsize=10)
    ax.grid(alpha=0.3); ax.legend(fontsize=7, ncol=2)
    return ax


def convergence_overlay(grid, function, models=None, acqs=None, n_rep=10, ground_truth=True,
                        ax=None):
    """Every (model, acquisition) on ONE regret axis, using the shared grammar: ACQUISITION -> colour,
    MODEL -> line style + marker. Two legends (colour = acq, style = model), so you can read both
    factors at once -- e.g. 'does the noise-aware acq (one colour) help every model (every style)?'."""
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D
    spec = P.get(function); fstar = P.ground_truth_min(spec)
    base = _true_best_traj if ground_truth else _noisy_best_traj
    models = models or _ordered_models(grid)
    acqs = acqs or list(acquisitions.CONFIG_ORDER)
    if ax is None:
        _, ax = plt.subplots(figsize=(8, 5))
    seen_a, seen_m = set(), set()
    for (a, p) in acqs:
        col = ACQ_COLORS.get(a, "C7")
        for m in models:
            ls, mk = MODEL_STYLES.get(m, ("-", None))
            runs = grid.select(function=function, model=m, acf=a, param=p, n_rep=n_rep)
            if not runs:
                continue
            trajs = [base(r, spec) - fstar for r in runs]
            L = min(len(t) for t in trajs)
            mean = np.array([t[:L] for t in trajs]).mean(0)
            ax.plot(np.arange(1, L + 1), np.maximum(mean, 1e-6), color=col, ls=ls, marker=mk,
                    markevery=max(1, L // 8), ms=4, lw=1.5, alpha=0.9)
            seen_a.add((a, p)); seen_m.add(m)
    ax.set_yscale("log"); ax.set_xlabel("BO iteration"); ax.set_ylabel("regret = value − f*")
    ax.set_title(f"{function}: all models × acquisitions (n_rep={n_rep})", fontsize=10)
    ax.grid(alpha=0.3, which="both")
    ah = [Line2D([0], [0], color=ACQ_COLORS.get(a, "C7"), lw=2.5, label=acquisitions.label(a, p))
          for (a, p) in acqs if (a, p) in seen_a]
    mh = [Line2D([0], [0], color="0.3", ls=MODEL_STYLES.get(m, ("-", None))[0],
                 marker=MODEL_STYLES.get(m, ("-", None))[1], label=_mlabel(m))
          for m in models if m in seen_m]
    leg1 = ax.legend(handles=ah, title="acquisition (colour)", fontsize=7, title_fontsize=7,
                     loc="upper right")
    ax.add_artist(leg1)
    ax.legend(handles=mh, title="model (style)", fontsize=7, title_fontsize=7, loc="lower left")
    return ax


def tradeoff_scatter(grid, function, models=None, acqs=None, n_rep=10, ground_truth=True, ax=None):
    """Final regret (y) vs σ² at the incumbent (x), one point per (model, acquisition). Lower-LEFT is
    the sweet spot: accurate optimum in a quiet region. Colour = acquisition, marker = model -- the
    core heteroscedastic trade-off in a single view."""
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D
    spec = P.get(function); fstar = P.ground_truth_min(spec)
    base = _true_best_traj if ground_truth else _noisy_best_traj
    models = models or _ordered_models(grid)
    acqs = acqs or list(acquisitions.CONFIG_ORDER)
    if ax is None:
        _, ax = plt.subplots(figsize=(7, 5))
    seen_a, seen_m = set(), set()
    for (a, p) in acqs:
        col = ACQ_COLORS.get(a, "C7")
        for m in models:
            runs = grid.select(function=function, model=m, acf=a, param=p, n_rep=n_rep)
            if not runs:
                continue
            reg = np.mean([base(r, spec)[-1] - fstar for r in runs])
            s2 = np.mean([sigma2_at_best_traj(r, spec, ground_truth)[-1] for r in runs])
            ax.scatter([s2], [max(reg, 1e-6)], color=col, marker=MODEL_STYLES.get(m, ("-", "o"))[1],
                       s=70, edgecolor="k", lw=0.5, alpha=0.9, zorder=3)
            seen_a.add((a, p)); seen_m.add(m)
    ax.set_yscale("log")
    ax.set_xlabel("σ² at incumbent (lower = quieter)"); ax.set_ylabel("final regret (lower = better)")
    ax.set_title(f"{function}: regret vs noise trade-off (↙ best)", fontsize=10)
    ax.grid(alpha=0.3, which="both")
    ah = [Line2D([0], [0], color=ACQ_COLORS.get(a, "C7"), lw=0, marker="o", label=acquisitions.label(a, p))
          for (a, p) in acqs if (a, p) in seen_a]
    mh = [Line2D([0], [0], color="0.3", lw=0, marker=MODEL_STYLES.get(m, ("-", "o"))[1], label=_mlabel(m))
          for m in models if m in seen_m]
    leg1 = ax.legend(handles=ah, title="acq (colour)", fontsize=6.5, title_fontsize=7, loc="upper right")
    ax.add_artist(leg1)
    ax.legend(handles=mh, title="model (marker)", fontsize=6.5, title_fontsize=7, loc="lower right")
    return ax


def sampling_trajectory(grid, function, acf="ei", param=float("nan"), n_rep=10, seed=1,
                        models=None, axes=None):
    """Where each model actually SAMPLED: x1 location (y) per BO iteration (x), coloured by the chosen
    categorical level; the dashed red line is the optimum's x1, the grey line the DOE→BO boundary.
    Reveals exploration vs exploitation and whether the model locks onto the right category.
    One panel per model (needs `axes` with >= len(models) entries)."""
    import matplotlib.pyplot as plt
    spec = P.get(function); lv_opt, x_opt = P.true_opt_location(spec)
    models = models or _ordered_models(grid)
    for ax, m in zip(np.ravel(axes), models):
        run = _one_run(grid, function, m, acf, param, n_rep, seed)
        if run is None:
            ax.text(0.5, 0.5, "no data yet", ha="center", va="center", transform=ax.transAxes,
                    color="crimson"); ax.set_title(_mlabel(m), fontsize=9); continue
        X = run["X_sampled"]; n0 = run["n_initial"]
        it = np.arange(len(X)) - n0 + 1                       # <=0 initial DOE, >0 BO iterations
        cols = [plt.cm.tab10((int(l) - 1) % 10) for l in X[:, -1]]
        ax.scatter(it, X[:, 0], c=cols, s=20, edgecolor="k", lw=0.2, zorder=3)
        ax.axhline(x_opt, color="crimson", ls="--", lw=1, zorder=2)
        ax.axvline(0.5, color="grey", ls=":", lw=1, zorder=2)
        ax.set_xlabel("BO iteration"); ax.set_ylabel("x1")
        ax.set_ylim(spec.lb, spec.ub)
        ax.set_title(f"{_mlabel(m)}  (final regret "
                     f"{_true_best_traj(run, spec)[-1] - P.ground_truth_min(spec):.3f})", fontsize=9)
        ax.grid(alpha=0.3)
    return axes


def acquisition_facet(grid, function, models=None, n_rep=10, ground_truth=True, ncol=3,
                      share_y=True, noise_aware_only=False, ax_size=(3.5, 3.5), center="mean"):#ax_size=(4.7, 3.5)):
    """One panel PER ACQUISITION (from CONFIG_ORDER, those with data), each overlaying every model's
    regret convergence (mean ± s.e., log axis). Colour = model. With `share_y` all panels use ONE
    y-range (from the mean curves) so the acquisitions are directly comparable at a glance.
    `noise_aware_only=True` restricts to the noise-aware acquisitions (HAEI/ANPEI/RAHBO) instead of all 6."""
    import matplotlib.pyplot as plt
    spec = P.get(function); fstar = P.ground_truth_min(spec)
    base = _true_best_traj if ground_truth else _noisy_best_traj
    models = models or _ordered_models(grid)
    cfgs = [(a, p) for (a, p) in acquisitions.CONFIG_ORDER
            if (not noise_aware_only or acquisitions.needs_aleatoric(a))
            and any(grid.select(function=function, model=m, acf=a, param=p, n_rep=n_rep) for m in models)]
    if not cfgs:
        fig, ax = plt.subplots(figsize=(5, 2)); ax.text(0.5, 0.5, "no data yet", ha="center",
            va="center", color="crimson", transform=ax.transAxes); ax.axis("off"); return fig
    nrow = (len(cfgs) + ncol - 1) // ncol
    fig, axes = plt.subplots(nrow, ncol, figsize=(ax_size[0] * ncol, ax_size[1] * nrow), squeeze=False)
    for ax, (a, p) in zip(axes.ravel(), cfgs):
        _panel(ax, grid, function, a, p, n_rep, lambda r, s: base(r, s) - fstar,
               "regret (true) = value − f*", True, center=center)
        ax.set_title(acquisitions.label(a, p), fontsize=10)
    for k in range(len(cfgs), nrow * ncol):
        axes.ravel()[k].axis("off")
    if share_y:                                          # ONE y-range from the mean lines (ignore the
        lo = hi = None                                   # ±s.e. band floor so a wide band can't squash it)
        for ax in axes.ravel()[:len(cfgs)]:
            for ln in ax.get_lines():
                yd = np.asarray(ln.get_ydata(), float); yd = yd[np.isfinite(yd) & (yd > 0)]
                if yd.size:
                    lo = yd.min() if lo is None else min(lo, yd.min())
                    hi = yd.max() if hi is None else max(hi, yd.max())
        if lo and hi:
            for ax in axes.ravel()[:len(cfgs)]:
                ax.set_ylim(lo * 0.7, hi * 1.4)
    tag = "true f" if ground_truth else "observed mean"
    fig.suptitle(f"{function} — regret per acquisition (n_rep={n_rep}, incumbent by {tag}"
                 f"{'; shared y' if share_y else ''})", y=1.003, fontsize=12)
    fig.tight_layout()
    return fig


def problem_report(grid, function, acf="ei", param=float("nan"), n_rep=10, ground_truth=True,
                   seed=1, save_dir=None, show=True, dpi=110, noise_aware_only=False):
    """Full single-problem report -> two figures:
      overview.png     landscape+DOE · regret · value+f* · σ² · all-model×acq overlay · regret-vs-noise
      exploration.png  the sampling trajectory of each model (x1 vs iteration, coloured by level)
    `acf`/`param` fix the single acquisition used in the per-model convergence + exploration panels;
    the overlay and trade-off panels sweep ALL acquisitions. Saves under save_dir/ if given."""
    import os as _os
    import matplotlib.pyplot as plt
    spec = P.get(function)
    tag = "true f" if ground_truth else "observed mean"

    fig = plt.figure(figsize=(19, 9))
    axl = fig.add_subplot(2, 3, 1); problem_landscape(function, grid, n_rep, seed, ax=axl)
    ax2 = fig.add_subplot(2, 3, 2); noise_per_level(function, ax=ax2)   # heteroscedastic noise σ per level
    # (regret for this acquisition is still in panel 5's all-model×acq overlay and in by_acquisition.png)
    ax3 = fig.add_subplot(2, 3, 3)
    ok = _panel(ax3, grid, function, acf, param, n_rep,
                _true_best_traj if ground_truth else _noisy_best_traj, "best value", False)
    if ok:
        fstar = P.ground_truth_min(spec); ax3.axhline(fstar, color="grey", ls="--", lw=1.2)
        ax3.annotate(f"f* = {fstar:.3f}", xy=(0.98, fstar), xycoords=("axes fraction", "data"),
                     ha="right", va="bottom", fontsize=7, color="grey")
    ax3.set_title(f"{function}: value — {acquisitions.label(acf, param)}", fontsize=10)
    ax4 = fig.add_subplot(2, 3, 4)
    _panel(ax4, grid, function, acf, param, n_rep,
           lambda r, s: sigma2_at_best_traj(r, s, ground_truth), "σ² at incumbent", True)
    ax4.set_title(f"{function}: σ² at incumbent — {acquisitions.label(acf, param)}", fontsize=10)
    convergence_overlay(grid, function, n_rep=n_rep, ground_truth=ground_truth,
                        ax=fig.add_subplot(2, 3, 5))
    tradeoff_scatter(grid, function, n_rep=n_rep, ground_truth=ground_truth,
                     ax=fig.add_subplot(2, 3, 6))
    fig.suptitle(f"{function} — per-problem report (n_rep={n_rep}, incumbent by {tag})",
                 fontsize=13, y=1.005)
    fig.tight_layout()

    models = _ordered_models(grid)
    ncol = 2; nrow = (len(models) + ncol - 1) // ncol
    fig2, axes2 = plt.subplots(nrow, ncol, figsize=(6.2 * ncol, 4.2 * nrow), squeeze=False)
    sampling_trajectory(grid, function, acf, param, n_rep, seed, models, axes2)
    for k in range(len(models), nrow * ncol):
        axes2.ravel()[k].axis("off")
    from matplotlib.lines import Line2D                          # legend mapping colour -> level
    lv_opt, _ = P.true_opt_location(spec)
    lv_handles = [Line2D([0], [0], color=plt.cm.tab10((lv - 1) % 10), lw=0, marker="o", ms=6,
                         label=f"level {lv}" + ("  (opt)" if lv == lv_opt else "")) for lv in spec.levels]
    lv_handles.append(Line2D([0], [0], color="crimson", ls="--", lw=1.2, label="optimum x1"))
    fig2.legend(handles=lv_handles, loc="lower center", ncol=len(lv_handles), fontsize=8,
                frameon=False, bbox_to_anchor=(0.5, -0.02))
    fig2.suptitle(f"{function} — sampling trajectory · {acquisitions.label(acf, param)}, seed {seed} "
                  f"(colour = chosen level)", fontsize=12, y=1.01)
    fig2.tight_layout(rect=(0, 0.03, 1, 1))

    fig3 = acquisition_facet(grid, function, models=models, n_rep=n_rep, ground_truth=ground_truth,
                             noise_aware_only=noise_aware_only)
    fig4 = latent_space(grid, function, acf=acf, param=param, n_rep=n_rep, seed=seed)

    if save_dir:
        _os.makedirs(save_dir, exist_ok=True)
        fig.savefig(_os.path.join(save_dir, "overview.png"), dpi=dpi, bbox_inches="tight")
        fig2.savefig(_os.path.join(save_dir, "exploration.png"), dpi=dpi, bbox_inches="tight")
        fig3.savefig(_os.path.join(save_dir, "by_acquisition.png"), dpi=dpi, bbox_inches="tight")
        fig4.savefig(_os.path.join(save_dir, "latent_space.png"), dpi=dpi, bbox_inches="tight")
    if not show:
        plt.close(fig); plt.close(fig2); plt.close(fig3); plt.close(fig4)
    return fig, fig2, fig3, fig4


def all_problem_reports(grid, functions=None, save_root="plots", acf="ei", param=float("nan"),
                        n_rep=10, ground_truth=True, seed=1, show=False, noise_aware_only=False):
    """Run problem_report for every problem into
    save_root/<problem>/{overview,exploration,by_acquisition}.png. Returns the dirs written.
    `show=False` keeps notebook output clean for a batch."""
    import os as _os
    fns = functions or grid.functions()
    out = []
    for fn in fns:
        d = _os.path.join(save_root, fn)
        problem_report(grid, fn, acf=acf, param=param, n_rep=n_rep, ground_truth=ground_truth,
                       seed=seed, save_dir=d, show=show, noise_aware_only=noise_aware_only)
        out.append(d)
        print(f"  wrote {d}/overview.png + exploration.png + by_acquisition.png + latent_space.png")
    return out


# ======================================================================================
#  LVGP LATENT SPACE vs GROUND TRUTH
#  LVGP maps each categorical level to a 2-D latent point z; levels that behave similarly
#  under the true objective should land close together. We compare the learned geometry to
#  the TRUE geometry (how different the levels' noise-free curves actually are).
#  Only the LVGP models expose z (matlab `hyper.z`); python models have no such embedding.
# ======================================================================================
LVGP_MODELS = ("standard_LVGP", "heter_LVGP")


def _level_curve_dist(spec, n=300):
    """Ground-truth distance between categorical levels = RMS gap between their noise-free curves
    f(x1│level) over the domain. This is the geometry a good latent embedding should reproduce."""
    x = np.linspace(spec.lb, spec.ub, n)
    F = np.array([spec.f_true_level(x, lv) for lv in spec.levels])       # (L, n)
    L = len(spec.levels)
    D = np.zeros((L, L))
    for i in range(L):
        for j in range(L):
            D[i, j] = np.sqrt(np.mean((F[i] - F[j]) ** 2))
    return D


def _latent_runs(grid, function, model, acf, param, n_rep):
    return [r for r in grid.runs if r["problem"] == function and r["model"] == model
            and r["acf"] == acf and r["n_rep"] == n_rep and r.get("hyper_z") is not None
            and (param != param or abs(r["param"] - param) < 1e-9)]


def latent_space(grid, function, models=None, acf="ei", param=float("nan"), n_rep=10, seed=1):
    """Compare the LVGP LATENT EMBEDDING of the categorical levels with the ground truth.
    Per LVGP model: (1) a latent map for one seed — each level's learned 2-D z, annotated, coloured by
    its true underlying value (spec.meta['cat_values']); (2) a recovery panel pooling ALL seeds —
    learned latent distance vs the true curve distance between level pairs, with a Spearman ρ (rank
    correlation, invariant to the rotation/reflection/scale ambiguity of the embedding). ρ→1 means the
    latent geometry reproduces the true category geometry."""
    import matplotlib.pyplot as plt
    from scipy.stats import spearmanr
    spec = P.get(function)
    L = spec.n_levels
    catv = np.asarray(spec.meta.get("cat_values", np.arange(1, L + 1)), float)
    lv_opt, _ = P.true_opt_location(spec)
    Dtrue = _level_curve_dist(spec)
    iu = np.triu_indices(L, k=1)
    dtrue = Dtrue[iu]
    models = [m for m in (models or LVGP_MODELS) if _latent_runs(grid, function, m, acf, param, n_rep)]
    if not models:
        fig, ax = plt.subplots(figsize=(5, 2)); ax.text(0.5, 0.5, "no LVGP latent data yet",
            ha="center", va="center", color="crimson", transform=ax.transAxes); ax.axis("off"); return fig

    ncol = len(models) + 1
    fig, axes = plt.subplots(1, ncol, figsize=(4.9 * ncol, 4.6), squeeze=False)
    axes = axes.ravel()
    for k, m in enumerate(models):
        runs = _latent_runs(grid, function, m, acf, param, n_rep)
        seed_run = next((r for r in runs if r["seed"] == seed), runs[0])
        z = np.asarray(seed_run["hyper_z"], float)
        ax = axes[k]
        sc = ax.scatter(z[:, 0], z[:, 1], c=catv, cmap="viridis", s=180, edgecolor="k", lw=0.6, zorder=3)
        for lv in range(L):
            ax.annotate(f"lv{lv+1}" + ("★" if lv + 1 == lv_opt else ""), (z[lv, 0], z[lv, 1]),
                        fontsize=8, ha="center", va="center",
                        color="w" if catv[lv] < catv.mean() else "k", zorder=4)
        cb = fig.colorbar(sc, ax=ax, fraction=0.046); cb.set_label("true level value", fontsize=8)
        ax.set_xlabel("latent z₁"); ax.set_ylabel("latent z₂")
        ax.set_title(f"{_mlabel(m)} — latent map (seed {seed_run['seed']})", fontsize=10)
        ax.grid(alpha=0.3); ax.set_aspect("equal", adjustable="datalim")

    axr = axes[-1]
    for m in models:
        runs = _latent_runs(grid, function, m, acf, param, n_rep)
        dz_pool, rhos = [], []
        for r in runs:
            z = np.asarray(r["hyper_z"], float)
            Dz = np.sqrt(((z[:, None, :] - z[None, :, :]) ** 2).sum(-1))[iu]
            if Dz.max() > 0:
                dz_pool.append(Dz / Dz.mean())               # scale-normalize per seed
                rhos.append(spearmanr(Dz, dtrue).statistic)
        if not dz_pool:
            continue
        dz_mean = np.mean(dz_pool, 0)
        col = MODEL_COLORS.get(m, "C7")
        axr.scatter(dtrue, dz_mean, color=col, s=45, edgecolor="k", lw=0.4, zorder=3,
                    label=f"{_mlabel(m)}  ρ={np.mean(rhos):.2f}±{np.std(rhos):.2f}")
    axr.set_xlabel("TRUE curve distance between levels")
    axr.set_ylabel("learned latent distance (seed-normalized)")
    axr.set_title(f"latent vs true geometry\n(Spearman ρ over {len(runs)} seeds; ρ→1 = faithful)",
                  fontsize=10)
    axr.grid(alpha=0.3); axr.legend(fontsize=8)
    fig.suptitle(f"{function} — LVGP latent space vs ground truth "
                 f"({acquisitions.label(acf, param)}, n_rep={n_rep})", y=1.02, fontsize=12)
    fig.tight_layout()
    return fig


# ======================================================================================
#  ACQUISITION-FUNCTION COMPARISON TABLES
# ======================================================================================
def acquisition_tables(grid, model="heter_LVGP", n_rep=10, functions=None):
    """Per-acquisition final performance for ONE model (rows = the 6 acquisitions, cols = problems).
    Returns (regret_df, noisy_df):
      regret = true_best − f*  (noise-free value at the sampled points; the honest metric)
      noisy  = best observed replicate-MEAN value (what the optimizer saw; can dip below f*)
    Both as 'mean ± sd' over seeds. Fixing the model isolates the acquisition effect; use a model that
    supports all 6 (heter_LVGP / separate_gp / categorical_kernel)."""
    import pandas as pd
    fns = functions or grid.functions()
    reg, noi, regn, noin = {}, {}, {}, {}
    for (a, p) in acquisitions.CONFIG_ORDER:
        lab = acquisitions.label(a, p); rrow, nrow, rn, nn = {}, {}, {}, {}
        for fn in fns:
            spec = P.get(fn); fstar = P.ground_truth_min(spec)
            runs = grid.select(function=fn, model=model, acf=a, param=p, n_rep=n_rep)
            if not runs:
                rrow[fn] = nrow[fn] = "—"; rn[fn] = nn[fn] = np.nan; continue
            r = np.array([_true_best_traj(x, spec)[-1] - fstar for x in runs])
            v = np.array([_noisy_best_traj(x, spec)[-1] for x in runs])
            rrow[fn] = f"{r.mean():.3g} ± {r.std():.2g}"; rn[fn] = float(r.mean())
            nrow[fn] = f"{v.mean():.3g} ± {v.std():.2g}"; nn[fn] = float(v.mean())
        reg[lab] = rrow; noi[lab] = nrow; regn[lab] = rn; noin[lab] = nn
    R = pd.DataFrame(reg).T[fns]; N = pd.DataFrame(noi).T[fns]
    R.index.name = N.index.name = f"{_mlabel(model)}: acquisition"
    R.attrs["means"] = pd.DataFrame(regn).T[fns]; N.attrs["means"] = pd.DataFrame(noin).T[fns]
    return R, N


def acquisition_ranking(grid, n_rep=10, functions=None):
    """CONCLUSIVE cross-problem ranking of the 6 acquisitions. Raw regret spans orders of magnitude
    across problems, so averaging it is meaningless -- this uses RANKS instead: within each
    (problem, model) cell the 6 acquisitions are ranked by mean final true regret (1 = best), then each
    acquisition's rank is averaged over every cell. Only models supporting all 6 acquisitions are used
    (heter_LVGP / separate_gp / categorical_kernel), so the comparison is on equal footing.
    Lower mean_rank = better overall; `wins` = # cells where it was the single best.
    Returns (dataframe, n_cells, models_used)."""
    import pandas as pd
    fns = functions or grid.functions()
    allacq = set(a for a, _ in acquisitions.CONFIG_ORDER)
    full = [m for m in _ordered_models(grid) if allacq <= set(MODELS[m].supports)]
    cfgs = list(acquisitions.CONFIG_ORDER)
    ranks = {acquisitions.label(a, p): [] for a, p in cfgs}
    wins = {acquisitions.label(a, p): 0 for a, p in cfgs}
    cells = 0
    for fn in fns:
        spec = P.get(fn); fstar = P.ground_truth_min(spec)
        for m in full:
            vals, complete = {}, True
            for (a, p) in cfgs:
                runs = grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
                if not runs:
                    complete = False; break
                vals[(a, p)] = np.mean([_true_best_traj(x, spec)[-1] - fstar for x in runs])
            if not complete:
                continue
            cells += 1
            order = sorted(cfgs, key=lambda c: vals[c])
            for rk, c in enumerate(order, 1):
                ranks[acquisitions.label(*c)].append(rk)
            wins[acquisitions.label(*order[0])] += 1
    rows = []
    for (a, p) in cfgs:
        lab = acquisitions.label(a, p); rr = ranks[lab]
        rows.append(dict(acquisition=lab, mean_rank=round(float(np.mean(rr)), 2) if rr else None,
                         median_rank=float(np.median(rr)) if rr else None,
                         wins=wins[lab], n_cells=len(rr)))
    df = pd.DataFrame(rows).set_index("acquisition").sort_values("mean_rank")
    return df, cells, [_mlabel(m) for m in full]


# ======================================================================================
#  METHOD (SURROGATE-MODEL) COMPARISON TABLES  +  the combined method×acq leaderboard
# ======================================================================================
def method_tables(grid, acf="ei", param=float("nan"), n_rep=10, functions=None):
    """Per-model final performance for ONE acquisition (rows = the 4 models, cols = problems).
    Default acf='ei' is noise-blind, so all 4 models (incl. standard_LVGP) appear. Returns
    (regret_df, noisy_df) as 'mean ± sd' over seeds -- same metrics as acquisition_tables."""
    import pandas as pd
    fns = functions or grid.functions()
    reg, noi, regn, noin = {}, {}, {}, {}
    for m in _ordered_models(grid):
        rrow, nrow, rn, nn = {}, {}, {}, {}
        for fn in fns:
            spec = P.get(fn); fstar = P.ground_truth_min(spec)
            runs = grid.select(function=fn, model=m, acf=acf, param=param, n_rep=n_rep)
            if not runs:
                rrow[fn] = nrow[fn] = "—"; rn[fn] = nn[fn] = np.nan; continue
            r = np.array([_true_best_traj(x, spec)[-1] - fstar for x in runs])
            v = np.array([_noisy_best_traj(x, spec)[-1] for x in runs])
            rrow[fn] = f"{r.mean():.3g} ± {r.std():.2g}"; rn[fn] = float(r.mean())
            nrow[fn] = f"{v.mean():.3g} ± {v.std():.2g}"; nn[fn] = float(v.mean())
        reg[_mlabel(m)] = rrow; noi[_mlabel(m)] = nrow; regn[_mlabel(m)] = rn; noin[_mlabel(m)] = nn
    R = pd.DataFrame(reg).T[fns]; N = pd.DataFrame(noi).T[fns]
    R.index.name = N.index.name = f"acq={acquisitions.label(acf, param)}: model"
    R.attrs["means"] = pd.DataFrame(regn).T[fns]; N.attrs["means"] = pd.DataFrame(noin).T[fns]
    return R, N


def _mean_regret(grid, fn, m, a, p, n_rep, spec, fstar):
    runs = grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
    return np.mean([_true_best_traj(x, spec)[-1] - fstar for x in runs]) if runs else None


def method_ranking(grid, n_rep=10, functions=None):
    """CONCLUSIVE model ranking, from two fair angles (raw regret can't be averaged across problems):
      mean_rank_bestacq : each model uses its OWN best acquisition per problem (min regret), then the
                          4 models are ranked within each problem and averaged -> 'model at its best'.
                          This is the fair view for heter_LVGP, whose edge needs a noise-aware acq.
      mean_rank_blind   : all 4 models ranked on the SHARED noise-blind acqs (ei/lcb/pi) only -- the
                          apples-to-apples same-acquisition comparison (12 cells).
    Lower = better. `wins_bestacq` = # problems the model's best config is the overall best."""
    import pandas as pd
    fns = functions or grid.functions()
    models = _ordered_models(grid)
    blind = [(a, p) for (a, p) in acquisitions.CONFIG_ORDER if not acquisitions.needs_aleatoric(a)]
    rb = {m: [] for m in models}; rl = {m: [] for m in models}; wins = {m: 0 for m in models}
    best_acq = {m: [] for m in models}
    for fn in fns:
        spec = P.get(fn); fstar = P.ground_truth_min(spec)
        # best-acq-per-model
        best = {}
        for m in models:
            cand = [(a, p) for (a, p) in acquisitions.CONFIG_ORDER if a in MODELS[m].supports]
            regs = [(_mean_regret(grid, fn, m, a, p, n_rep, spec, fstar), (a, p)) for (a, p) in cand]
            regs = [(r, ap) for r, ap in regs if r is not None]
            if regs:
                r, ap = min(regs, key=lambda t: t[0]); best[m] = r; best_acq[m].append(acquisitions.label(*ap))
        order = sorted(best, key=lambda m: best[m])
        for rk, m in enumerate(order, 1):
            rb[m].append(rk)
        if order:
            wins[order[0]] += 1
        # shared noise-blind ranking, per blind acq
        for (a, p) in blind:
            vals = {m: _mean_regret(grid, fn, m, a, p, n_rep, spec, fstar) for m in models}
            vals = {m: v for m, v in vals.items() if v is not None}
            for rk, m in enumerate(sorted(vals, key=lambda m: vals[m]), 1):
                rl[m].append(rk)
    rows = []
    from collections import Counter
    for m in models:
        rows.append(dict(model=_mlabel(m),
                         mean_rank_bestacq=round(float(np.mean(rb[m])), 2) if rb[m] else None,
                         wins_bestacq=wins[m],
                         mean_rank_blind=round(float(np.mean(rl[m])), 2) if rl[m] else None,
                         usual_best_acq=Counter(best_acq[m]).most_common(1)[0][0] if best_acq[m] else "—"))
    return pd.DataFrame(rows).set_index("model").sort_values("mean_rank_bestacq")


def method_acq_leaderboard(grid, n_rep=10, functions=None, top=None):
    """THE combined leaderboard: every (model × acquisition) configuration ranked. Within each problem
    all available configs are ranked by mean regret (1=best); mean_rank is averaged over the problems.
    Top row = the single best configuration overall. `wins` = # problems it is the outright best config.
    This answers 'what is the best method+acquisition to use', pooling both factors. `top` truncates."""
    import pandas as pd
    fns = functions or grid.functions()
    combos = [(m, a, p) for m in _ordered_models(grid)
              for (a, p) in acquisitions.CONFIG_ORDER if a in MODELS[m].supports]
    ranks = {c: [] for c in combos}; wins = {c: 0 for c in combos}
    for fn in fns:
        spec = P.get(fn); fstar = P.ground_truth_min(spec)
        vals = {c: _mean_regret(grid, fn, c[0], c[1], c[2], n_rep, spec, fstar) for c in combos}
        avail = [c for c in combos if vals[c] is not None]
        order = sorted(avail, key=lambda c: vals[c])
        for rk, c in enumerate(order, 1):
            ranks[c].append(rk)
        if order:
            wins[order[0]] += 1
    rows = []
    for (m, a, p) in combos:
        rr = ranks[(m, a, p)]
        rows.append(dict(model=_mlabel(m), acquisition=acquisitions.label(a, p),
                         mean_rank=round(float(np.mean(rr)), 2) if rr else None,
                         best_rank=int(min(rr)) if rr else None, wins=wins[(m, a, p)]))
    df = pd.DataFrame(rows).sort_values("mean_rank").set_index(["model", "acquisition"])
    return df if top is None else df.head(top)


# Risk-aversion coefficient for the mean-variance (RAHBO) robustness metric MV = f + alpha*sigma^2.
# alpha is a modelling choice (how much you penalise noise); rankings can shift with it.
MV_ALPHA = 1.0
_MV_STAR = {}


def _mv_star(spec, alpha, n=4000):
    """True mean-variance optimum  min_{x,ℓ} [ f(x,ℓ) + alpha*sigma(x,ℓ)^2 ]  (cached per (problem, alpha))."""
    key = (spec.name, alpha)
    if key not in _MV_STAR:
        x = np.linspace(spec.lb, spec.ub, n)
        _MV_STAR[key] = min(float((spec.f_true_level(x, lv) + alpha * spec.sigma_level(x, lv) ** 2).min())
                            for lv in spec.levels)
    return _MV_STAR[key]


def _mv_sampled(run, spec, alpha):
    """True MV = f + alpha*sigma^2 at every sampled point (from the stored true f/sigma; recomputed on a v1 cell)."""
    f = run.get("f_true_sampled"); sig = run.get("sigma_true_sampled")
    if f is None or sig is None:
        X = run["X_sampled"]
        f = np.array([float(np.ravel(spec.f_true_level(x[:-1], int(round(x[-1]))))[0]) for x in X])
        sig = np.array([float(np.ravel(spec.sigma_level(x[:-1], int(round(x[-1]))))[0]) for x in X])
    return np.asarray(f, float) + alpha * np.asarray(sig, float) ** 2


def _final_metric(run, spec, fstar, metric, ground_truth):
    """Final value of the chosen metric for one run (all lower-is-better):
      'regret' value − f* · 'noise' σ² at incumbent · 'mv' mean-variance regret (RAHBO robustness)."""
    if metric == "regret":
        base = _true_best_traj if ground_truth else _noisy_best_traj
        return base(run, spec)[-1] - fstar
    if metric == "noise":                                    # true σ² at the incumbent best design
        return sigma2_at_best_traj(run, spec, ground_truth)[-1]
    if metric == "mv":                                       # RAHBO mean-variance regret: robust optimum
        return float(np.min(_mv_sampled(run, spec, MV_ALPHA))) - _mv_star(spec, MV_ALPHA)
    raise ValueError(f"metric must be 'regret', 'noise' or 'mv', got {metric!r}")


def acq_method_table(grid, metric="regret", n_rep=10, functions=None, ground_truth=True):
    """Full table over EVERY (acquisition × model) configuration for one METRIC (both lower-is-better):
      metric='regret' -> final true regret (value − f*)
      metric='noise'  -> final σ² at the incumbent best design (the noise you'd deploy into)
    rows = MultiIndex (acquisition, model), cols = problems, cells = 'mean ± sd'. Includes every model
    that supports each acquisition (standard_LVGP appears only on the noise-blind acquisitions).
    Style two ways: style_best(df, group_level='acquisition') = best model per acquisition;
    style_best(df) = single best (acq, model) combo per problem."""
    import pandas as pd
    fns = functions or grid.functions()
    rows, nums = {}, {}
    for (a, p) in acquisitions.CONFIG_ORDER:
        for m in _ordered_models(grid):
            if a not in MODELS[m].supports:
                continue
            row, nrow = {}, {}
            for fn in fns:
                spec = P.get(fn); fstar = P.ground_truth_min(spec)
                runs = grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
                if not runs:
                    row[fn] = "—"; nrow[fn] = np.nan; continue
                v = np.array([_final_metric(x, spec, fstar, metric, ground_truth) for x in runs])
                row[fn] = f"{v.mean():.3g} ± {v.std():.2g}"; nrow[fn] = float(v.mean())
            key = (acquisitions.label(a, p), _mlabel(m))
            rows[key] = row; nums[key] = nrow
    df = pd.DataFrame(rows).T[fns]
    df.index = pd.MultiIndex.from_tuples(df.index, names=["acquisition", "model"])
    nd = pd.DataFrame(nums).T[fns]; nd.index = df.index
    df.attrs["means"] = nd                                    # full-precision means for tie-free highlighting
    return df


def acq_method_regret_table(grid, n_rep=10, functions=None):
    """Backwards-compatible wrapper: the regret variant of acq_method_table."""
    return acq_method_table(grid, "regret", n_rep, functions, ground_truth=True)


def acq_method_heatmap(grid, metric="regret", n_rep=10, functions=None, ground_truth=True, ncol=2):
    """Conventional per-problem heatmaps (rows = acquisition, cols = model, viridis value colorbar) for
    the given metric ('regret' or 'noise'). Thin wrapper over heatmaps_by_function."""
    return heatmaps_by_function(grid, functions=functions, n_rep=n_rep, ground_truth=ground_truth,
                                ncol=ncol, metric=metric)


def export_acq_method_excel(grid, path="acq_method_tables.xlsx", n_reps=(3, 10), functions=None,
                            ground_truth=True):
    """SEPARATE workbook for the full (acquisition × method) tables: ONE combined sheet per
    {regret, noise} × {n_reps} -- bold = best model per acquisition, green = best combo overall.
    (One sheet instead of the old per-acq + combo pair.)"""
    fns = functions or grid.functions()
    specs = []
    for metric in ("regret", "noise", "mv"):                 # mv = RAHBO mean-variance robustness
        for nr in n_reps:
            df = acq_method_table(grid, metric, nr, fns, ground_truth)
            colmin = {c: "min" for c in df.columns}
            specs.append((f"{metric}_nrep{nr:02d}", df, colmin, "acquisition", "combined"))
    export_tables_excel(specs, path)
    return path


# ======================================================================================
#  TABLE FORMATTING: bold-the-best styling (notebook) + formatted multi-sheet Excel export
# ======================================================================================
def _cell_mean(x):
    """Leading number of a 'mean ± sd' string (or a plain number); NaN if unparseable/'—'."""
    try:
        return float(str(x).split("±")[0])
    except Exception:
        return float("nan")


_BEST_CSS = "font-weight:bold;background-color:#c6efce"


def _best_mask(df, directions=None, group_level=None):
    """Boolean array (rows × cols): True where a cell is the best in its column. group_level=None -> best
    over the WHOLE column; an index level -> best WITHIN each group of that level (e.g. per acquisition).
    Comparison uses the FULL-PRECISION means in df.attrs['means'] when present (so cells that merely
    ROUND to the same displayed value, e.g. many '0.01' noise cells, don't all tie); otherwise it falls
    back to parsing the displayed 'mean ± sd' string. Constant / all-NaN columns/groups get no marks."""
    import pandas as pd
    means = getattr(df, "attrs", {}).get("means")
    if means is not None:
        num = means.reindex(index=df.index, columns=df.columns)
    else:
        num = df.apply(lambda c: c.map(_cell_mean))
    mask = np.zeros(df.shape, bool)
    lvl = np.asarray(df.index.get_level_values(group_level)) if group_level is not None else None
    for jc, col in enumerate(df.columns):
        v = num[col].to_numpy(float)
        d = (directions or {}).get(col, "min")
        groups = [np.arange(len(v))] if lvl is None else [np.where(lvl == g)[0] for g in pd.unique(lvl)]
        for pos in groups:
            sub = v[pos]
            if np.all(np.isnan(sub)) or np.nanmin(sub) == np.nanmax(sub):
                continue
            best = np.nanmin(sub) if d == "min" else np.nanmax(sub)
            mask[pos[sub == best], jc] = True
    return mask


def style_best(df, directions=None, group_level=None):
    """pandas Styler that BOLDS + green-fills the best cell per column. `directions` maps a column to
    'min' (default) or 'max'. `group_level` (an index level) restricts "best" to WITHIN each group of
    that level -- e.g. group_level='acquisition' bolds the best model for each acquisition."""
    import pandas as pd
    m = _best_mask(df, directions, group_level)
    styles = pd.DataFrame(np.where(m, _BEST_CSS, ""), index=df.index, columns=df.columns)
    return df.style.apply(lambda _: styles, axis=None)


def style_best_combined(df, directions=None, group_level="acquisition"):
    """ONE table, two levels of emphasis (so two separate tables aren't needed):
      **bold**            = best cell WITHIN each `group_level` group per column (e.g. best model per
                            acquisition),
      **bold + green**    = best cell over the WHOLE column (the single best (acq × model) combo).
    The overall best is also its group's best, so it gets both."""
    import pandas as pd
    mg = _best_mask(df, directions, group_level)             # best within each group
    mo = _best_mask(df, directions, None)                    # best overall (per column)
    css = np.where(mo, _BEST_CSS, np.where(mg, "font-weight:bold", ""))
    styles = pd.DataFrame(css, index=df.index, columns=df.columns)
    return df.style.apply(lambda _: styles, axis=None)


def export_tables_excel(specs, path):
    """Write formatted tables to a multi-sheet .xlsx. `specs` entries:
      (sheet, df, directions)                          -> best cell per column bold + green
      (sheet, df, directions, group_level)             -> best WITHIN each group bold + green
      (sheet, df, directions, group_level, 'combined') -> best-per-group BOLD, best-overall bold + GREEN
    Header bold + frozen; auto-width. Returns the path."""
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    fill = PatternFill("solid", fgColor="C6EFCE")
    with pd.ExcelWriter(path, engine="openpyxl") as xw:
        for spec in specs:
            sheet, df, directions = spec[0], spec[1], spec[2]
            group_level = spec[3] if len(spec) > 3 else None
            combined = len(spec) > 4 and spec[4] == "combined"
            sh = sheet[:31]
            df.to_excel(xw, sheet_name=sh)
            ws = xw.sheets[sh]
            nidx = df.index.nlevels
            for cell in ws[1]:
                cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="center")
            if combined:
                for i, jc in zip(*np.where(_best_mask(df, directions, group_level))):
                    ws.cell(row=i + 2, column=nidx + jc + 1).font = Font(bold=True)
                for i, jc in zip(*np.where(_best_mask(df, directions, None))):
                    cc = ws.cell(row=i + 2, column=nidx + jc + 1)
                    cc.font = Font(bold=True); cc.fill = fill
            else:
                for i, jc in zip(*np.where(_best_mask(df, directions, group_level))):
                    cc = ws.cell(row=i + 2, column=nidx + jc + 1)
                    cc.font = Font(bold=True); cc.fill = fill
            for k in range(1, ws.max_column + 1):
                w = max((len(str(ws.cell(row=r, column=k).value))
                         for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=k).value is not None),
                        default=10)
                ws.column_dimensions[get_column_letter(k)].width = min(max(w + 2, 11), 42)
            ws.freeze_panes = ws.cell(row=2, column=nidx + 1)
    return path


def export_comparison_tables(grid, path="comparison_tables.xlsx", n_rep=10, functions=None,
                             acq_model="heter_LVGP"):
    """Build every acquisition/method/leaderboard table and write them to ONE formatted .xlsx (best
    cells bold). Returns (path, specs) so the same specs can be reused for styled notebook display."""
    fns = functions or grid.functions()
    aR, aN = acquisition_tables(grid, model=acq_model, n_rep=n_rep, functions=fns)
    aRank, _, _ = acquisition_ranking(grid, n_rep=n_rep, functions=fns)
    mR, mN = method_tables(grid, acf="ei", n_rep=n_rep, functions=fns)
    mRank = method_ranking(grid, n_rep=n_rep, functions=fns)
    lb = method_acq_leaderboard(grid, n_rep=n_rep, functions=fns)
    colmin = {c: "min" for c in fns}
    specs = [
        (f"acq_regret_{acq_model}", aR, colmin),
        ("acq_noisy_value", aN, colmin),
        ("acq_ranking", aRank, {"mean_rank": "min", "median_rank": "min", "wins": "max"}),
        # (the full acquisition × method tables live in the separate acq_method_tables.xlsx)
        ("method_regret_EI", mR, colmin),
        ("method_noisy_value_EI", mN, colmin),
        ("method_ranking", mRank, {"mean_rank_bestacq": "min", "mean_rank_blind": "min",
                                   "wins_bestacq": "max"}),
        ("leaderboard", lb, {"mean_rank": "min", "best_rank": "min", "wins": "max"}),
    ]
    export_tables_excel(specs, path)
    return path, specs


# ======================================================================================
#  LaTeX export of the (acquisition × method) tables — booktabs, combined highlighting
# ======================================================================================
_PRETTY_FN = {
    "branin_hetero": "Branin", "sixhump_camel": "Six-hump Camel", "griewank_2d": "Griewank 2-D",
    "ackley_2d": "Ackley 2-D", "griewank_10d": "Griewank 10-D", "ackley_10d": "Ackley 10-D",
    "rastrigin_6d": "Rastrigin 6-D", "golinski": "Golinski", "piston": "Piston", "otl_circuit": "OTL Circuit",
}


def _sci_to_tex(t):
    """'2.6e-08' -> '2.6{\\times}10^{-8}'; plain numbers pass through."""
    import re
    t = t.strip()
    m = re.fullmatch(r"(-?\d*\.?\d+)[eE]([+-]?\d+)", t)
    return f"{m.group(1)}{{\\times}}10^{{{int(m.group(2))}}}" if m else t


def _acq_tex(lab):
    """'HAEI(γ=0.5)' -> 'HAEI ($\\gamma{=}0.5$)'; greek letters into math."""
    g = {"γ": "\\gamma", "β": "\\beta", "α": "\\alpha"}
    if "(" in lab:
        name, rest = lab.split("(", 1)
        rest = rest.rstrip(")")
        for k, v in g.items():
            rest = rest.replace(k, v)
        return f"{name} (${rest}$)"
    return lab


def _tex_cell(raw, bold, green):
    """One 'mean ± sd' cell -> LaTeX math. {\\boldmath} if best-in-group; \\colorbox (green) if
    best-overall. \\colorbox is used (not \\cellcolor) so only xcolor is needed -- no colortbl/multirow."""
    s = str(raw).strip()
    if s in ("—", "-", "nan", "None", ""):
        body = "--"
    else:
        parts = s.split("±")
        mean = _sci_to_tex(parts[0].strip())
        body = f"${mean} \\pm {_sci_to_tex(parts[1].strip())}$" if len(parts) > 1 else f"${mean}$"
    if bold:
        body = "{\\boldmath" + body + "}"
    if green:
        body = "\\colorbox{green!20}{" + body + "}"
    return body


def to_latex_combined(df, directions=None, group_level="acquisition", caption="", label=""):
    """LaTeX table (booktabs) for a MultiIndex (acquisition, model) DataFrame, with the combined
    highlighting: BOLD = best model within each acquisition (per problem), GREEN = best (acq, model)
    combo per problem. Acquisitions are grouped and separated by \\midrule (label on the first row of
    each group -- no multirow dependency)."""
    mg = _best_mask(df, directions, group_level)
    mo = _best_mask(df, directions, None)
    fns = list(df.columns)
    idx = list(df.index)
    order = []
    for a, _ in idx:
        if a not in order:
            order.append(a)
    L = ["\\begin{table}[htbp]", "\\centering", "\\small",
         "\\setlength{\\tabcolsep}{4pt}", "\\setlength{\\fboxsep}{2pt}"]
    if caption:
        L.append("\\caption{%s}" % caption)
    if label:
        L.append("\\label{%s}" % label)
    L.append("\\begin{tabular}{ll" + "r" * len(fns) + "}")
    L.append("\\toprule")
    L.append("Acquisition & Model & " +
             " & ".join(_PRETTY_FN.get(c, c.replace("_", "\\_")) for c in fns) + " \\\\")
    L.append("\\midrule")
    for gi, a in enumerate(order):
        krows = [k for k, (aa, _) in enumerate(idx) if aa == a]
        for j, k in enumerate(krows):
            acq_cell = _acq_tex(a) if j == 0 else ""            # label on first row of the group
            cells = " & ".join(_tex_cell(df.iloc[k, jc], mg[k, jc], mo[k, jc]) for jc in range(len(fns)))
            L.append("%s & %s & %s \\\\" % (acq_cell, idx[k][1], cells))
        L.append("\\midrule" if gi < len(order) - 1 else "\\bottomrule")
    L += ["\\end{tabular}", "\\end{table}"]
    return "\n".join(L)


_LATEX_PREAMBLE = r"""\documentclass[11pt]{article}
\usepackage[margin=0.8in,landscape]{geometry}
\usepackage{booktabs}
\usepackage[dvipsnames]{xcolor}
\usepackage{amsmath}
\title{Benchmark comparison tables --- acquisition $\times$ surrogate model}
\author{}
\date{}
\begin{document}
\maketitle
\noindent\textbf{Legend.} Each cell is the final metric (mean $\pm$ sd over seeds). \textbf{Bold} marks
the best model for that acquisition; \colorbox{green!20}{green} marks the single best (acquisition,
model) combination for that problem. Lower is better for both metrics.
\vspace{1em}
"""


def export_latex_tables(grid, out_dir="latex", n_reps=(3, 10), functions=None, ground_truth=True,
                        metrics=("regret", "noise", "mv"), compile_pdf=False):
    """Write one .tex table per {metric} × {n_rep} (same data as acq_method_tables.xlsx) plus a
    main.tex that \\input's them all. Returns the list of files written. compile_pdf tries pdflatex."""
    import os
    import subprocess
    os.makedirs(out_dir, exist_ok=True)
    fns = functions or grid.functions()
    desc = {"regret": "Final regret (value $-$ $f^*$)",
            "noise": "Final noise $\\sigma^2$ at the incumbent design",
            "mv": ("Final mean-variance regret (RAHBO robustness, "
                   "$\\mathrm{MV}=f+\\alpha\\sigma^2$, $\\alpha=%g$)" % MV_ALPHA)}
    written, inputs = [], []
    for metric in metrics:
        for nr in n_reps:
            df = acq_method_table(grid, metric, nr, fns, ground_truth)
            cap = ("%s for every acquisition $\\times$ surrogate model "
                   "($n_{\\mathrm{rep}}=%d$, mean $\\pm$ sd over seeds)." % (desc[metric], nr))
            tex = to_latex_combined(df, {c: "min" for c in df.columns}, "acquisition",
                                    caption=cap, label="tab:%s_nrep%02d" % (metric, nr))
            f = os.path.join(out_dir, "%s_nrep%02d.tex" % (metric, nr))
            with open(f, "w") as fh:
                fh.write(tex + "\n")
            written.append(f); inputs.append(os.path.basename(f))
    main = os.path.join(out_dir, "main.tex")
    with open(main, "w") as fh:
        fh.write(_LATEX_PREAMBLE + "\n".join("\\input{%s}\n\\clearpage" % i for i in inputs)
                 + "\n\\end{document}\n")
    written.append(main)
    if compile_pdf:
        subprocess.run(["pdflatex", "-interaction=nonstopmode", "main.tex"], cwd=out_dir,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        subprocess.run(["pdflatex", "-interaction=nonstopmode", "main.tex"], cwd=out_dir,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        pdf = os.path.join(out_dir, "main.pdf")
        if os.path.exists(pdf):
            written.append(pdf)
    return written


# ======================================================================================
#  PER-SEED DISTRIBUTION VIEWS
#  Summaries (mean ± sd) hide skew: on branin, RAHBO's MV-regret has mean 0.0040 but MEDIAN 0.0002 --
#  a few bad seeds inflate the mean 20x. These views show the whole per-seed distribution, so a
#  "typically much better but occasionally fails" method cannot be mistaken for a tie.
# ======================================================================================
def _seed_vals(grid, function, model, acf, param, n_rep, metric, ground_truth):
    spec = P.get(function); fstar = P.ground_truth_min(spec)
    runs = grid.select(function=function, model=model, acf=acf, param=param, n_rep=n_rep)
    return np.array([_final_metric(r, spec, fstar, metric, ground_truth) for r in runs])


def seed_boxplot(grid, function, metric="mv", n_rep=10, models=None, ground_truth=True,
                 logy=True, floor=1e-8, show_points=True, ax=None):
    """Per-seed distribution of the FINAL metric for every (acquisition × model) config on one problem.
    Box = median + IQR (+ whiskers); dots = the individual seeds. Grouped by acquisition, coloured by
    model. Log y (values span orders of magnitude; non-positive values are clipped to `floor`).
    This is the view that exposes skew -- a low median with a few bad outliers looks nothing like its mean."""
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D
    models = models or _ordered_models(grid)
    cfgs = list(acquisitions.CONFIG_ORDER)
    data, pos, cols, xt, xl = [], [], [], [], []
    rng = np.random.default_rng(0)
    for gi, (a, p) in enumerate(cfgs):
        present = [m for m in models if a in MODELS[m].supports
                   and len(grid.select(function=function, model=m, acf=a, param=p, n_rep=n_rep))]
        if not present:
            continue
        k = len(present)
        for mi, m in enumerate(present):
            v = _seed_vals(grid, function, m, a, p, n_rep, metric, ground_truth)
            data.append(np.maximum(v, floor) if logy else v)
            pos.append(gi + (mi - (k - 1) / 2) * (0.8 / k))
            cols.append(MODEL_COLORS.get(m, "C7"))
        xt.append(gi); xl.append(acquisitions.label(a, p))
    if not data:
        if ax is None:
            _, ax = plt.subplots()
        ax.text(0.5, 0.5, "no data", ha="center", va="center", transform=ax.transAxes); return ax
    if ax is None:
        _, ax = plt.subplots(figsize=(1.9 * len(xt) + 3, 5))
    w = 0.8 / max(len(models), 1) * 0.85
    bp = ax.boxplot(data, positions=pos, widths=w, patch_artist=True, showfliers=False,
                    medianprops=dict(color="k", lw=1.6))
    for patch, c in zip(bp["boxes"], cols):
        patch.set_facecolor(c); patch.set_alpha(0.5)
    if show_points:
        for v, pp, c in zip(data, pos, cols):
            ax.scatter(pp + rng.uniform(-w * 0.25, w * 0.25, len(v)), v, s=7, color=c,
                       edgecolor="k", lw=0.2, alpha=0.7, zorder=3)
    if logy:
        ax.set_yscale("log")
    ax.set_xticks(xt); ax.set_xticklabels(xl, rotation=20, ha="right", fontsize=8)
    lab = {"regret": "final regret", "noise": "σ² at incumbent", "mv": "MV-regret (robustness)"}[metric]
    ax.set_ylabel(lab + (" (log)" if logy else ""))
    ax.set_title(f"{function} — per-seed distribution of {lab} (n_rep={n_rep})\n"
                 f"box = median + IQR · dots = individual seeds", fontsize=10)
    ax.grid(alpha=0.3, axis="y", which="both")
    ax.legend(handles=[Line2D([0], [0], marker="s", lw=0, markerfacecolor=MODEL_COLORS.get(m, "C7"),
                              markeredgecolor="k", markersize=8, label=_mlabel(m)) for m in models],
              fontsize=8, loc="best")
    return ax


def ecdf(grid, function, configs, metric="mv", n_rep=10, ground_truth=True, logx=True, floor=1e-8,
         ax=None):
    """Empirical CDF of the final metric across seeds, for a chosen list of configs
    [(model, acf, param), ...]. Reads as: 'what fraction of seeds achieved a metric <= x'.
    A curve that rises FAR LEFT is typically better; a long flat tail on the right = occasional failures.
    This separates 'usually great, sometimes fails' (RAHBO) from 'consistently mediocre' -- which the
    mean cannot."""
    import matplotlib.pyplot as plt
    if ax is None:
        _, ax = plt.subplots(figsize=(7.5, 5))
    for i, (m, a, p) in enumerate(configs):
        v = _seed_vals(grid, function, m, a, p, n_rep, metric, ground_truth)
        if not len(v):
            continue
        v = np.sort(np.maximum(v, floor) if logx else np.sort(v))
        y = np.arange(1, len(v) + 1) / len(v)
        ls, mk = MODEL_STYLES.get(m, ("-", None))
        ax.step(v, y, where="post", lw=2, ls=ls, color=ACQ_COLORS.get(a, f"C{i}"),
                label=f"{_mlabel(m)} + {acquisitions.label(a, p)}  (med={np.median(v):.2g})")
        ax.axvline(np.median(v), color=ACQ_COLORS.get(a, f"C{i}"), ls=":", lw=1, alpha=0.5)
    if logx:
        ax.set_xscale("log")
    lab = {"regret": "final regret", "noise": "σ² at incumbent", "mv": "MV-regret"}[metric]
    ax.set_xlabel(lab + " (log)"); ax.set_ylabel("fraction of seeds ≤ x")
    ax.set_title(f"{function} — ECDF over seeds (n_rep={n_rep})\ndotted = median; left & steep = better",
                 fontsize=10)
    ax.grid(alpha=0.3, which="both"); ax.legend(fontsize=8, loc="lower right"); ax.set_ylim(0, 1.02)
    return ax


def head_to_head(grid, function, configs, metric="mv", n_rep=10, ground_truth=True):
    """Mean vs MEDIAN + a rank test for a shortlist of configs -- so a skew-driven 'tie' on the mean
    can't hide a real difference. Returns a DataFrame; also reports pairwise Mann-Whitney vs the row
    with the best median."""
    import pandas as pd
    from scipy import stats
    rows, vals = [], {}
    for (m, a, p) in configs:
        v = _seed_vals(grid, function, m, a, p, n_rep, metric, ground_truth)
        if not len(v):
            continue
        key = f"{_mlabel(m)} + {acquisitions.label(a, p)}"
        vals[key] = v
        rows.append(dict(config=key, mean=v.mean(), median=float(np.median(v)),
                         iqr=float(np.percentile(v, 75) - np.percentile(v, 25)),
                         worst=v.max(), n=len(v)))
    df = pd.DataFrame(rows).set_index("config").sort_values("median")
    best = df.index[0]
    df["p_vs_best_median"] = [
        "—" if k == best else f"{stats.mannwhitneyu(vals[k], vals[best], alternative='greater').pvalue:.1e}"
        for k in df.index]
    return df.round(6)


def mv_alpha_sweep(grid, alphas=(1, 5, 15), n_rep=10, functions=None):
    """Sensitivity of the robustness verdict to the risk-aversion alpha in MV = f + alpha*sigma^2.
    For each alpha: the TRUE robust optimum (level, x1) per problem, the best (model x acq) config by
    mean MV-regret, and the acquisition mean-rank over the full-support models. The 'which problems
    test robustness' answer is alpha-dependent: griewank_2d's best level flips 1->2 above
    alpha ~ 12.9 and branin's flips 2->4 above ~ 7.3 (see the per-level tension table)."""
    import pandas as pd
    global MV_ALPHA
    fns = functions or grid.functions()
    full = [m for m in _ordered_models(grid)
            if set(a for a, _ in acquisitions.CONFIG_ORDER) <= set(MODELS[m].supports)]
    keep = MV_ALPHA
    rows, rank_rows = [], []
    try:
        for alpha in alphas:
            MV_ALPHA = float(alpha)
            ranks = {acquisitions.label(a, p): [] for a, p in acquisitions.CONFIG_ORDER}
            for fn in fns:
                spec = P.get(fn); fstar = P.ground_truth_min(spec)
                x = np.linspace(spec.lb, spec.ub, 6001); best_star = (np.inf, None, None)
                for lv in spec.levels:
                    mv = spec.f_true_level(x, lv) + alpha * spec.sigma_level(x, lv) ** 2
                    i = int(mv.argmin())
                    if mv[i] < best_star[0]:
                        best_star = (float(mv[i]), lv, float(x[i]))
                best = (np.inf, "")
                for m in _ordered_models(grid):
                    for a, p in acquisitions.CONFIG_ORDER:
                        if a not in MODELS[m].supports:
                            continue
                        runs = grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
                        if not runs:
                            continue
                        v = np.mean([_final_metric(r, spec, fstar, "mv", True) for r in runs])
                        if v < best[0]:
                            best = (v, f"{_mlabel(m)} + {acquisitions.label(a, p)}")
                        if m in full:
                            ranks.setdefault(acquisitions.label(a, p), [])
                rows.append(dict(alpha=alpha, problem=fn, mv_star_level=best_star[1],
                                 mv_star_x1=round(best_star[2], 3),
                                 best_config=best[1], best_mv_regret=round(best[0], 4)))
                for m in full:
                    vals = {}
                    for a, p in acquisitions.CONFIG_ORDER:
                        runs = grid.select(function=fn, model=m, acf=a, param=p, n_rep=n_rep)
                        if runs:
                            vals[(a, p)] = np.mean([_final_metric(r, spec, fstar, "mv", True)
                                                    for r in runs])
                    for rk, (a, p) in enumerate(sorted(vals, key=lambda c: vals[c]), 1):
                        ranks[acquisitions.label(a, p)].append(rk)
            for lab, rr in ranks.items():
                if rr:
                    rank_rows.append(dict(alpha=alpha, acquisition=lab,
                                          mean_rank=round(float(np.mean(rr)), 2)))
    finally:
        MV_ALPHA = keep
    per_problem = pd.DataFrame(rows).set_index(["alpha", "problem"])
    acq_ranks = (pd.DataFrame(rank_rows).pivot(index="acquisition", columns="alpha",
                                               values="mean_rank"))
    return per_problem, acq_ranks
